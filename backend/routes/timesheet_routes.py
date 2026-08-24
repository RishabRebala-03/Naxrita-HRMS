# routes/timesheet_routes.py
from flask import Blueprint, request, jsonify, send_file
from bson import ObjectId
from datetime import datetime, timedelta
from io import BytesIO
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config.db import mongo
from services.queue_service import enqueue_mail
from services.smtp_service import resolve_tenant_id
from utils.access_control import has_admin_menu_access, require_admin_menu_access, resolve_requester, require_admin
from utils.timezone import now_ist

timesheet_bp = Blueprint("timesheet_bp", __name__)
WORKDAY_HOURS = 9.0
TIMESHEET_NOTIFICATION_REMINDER_HOURS = 24
TIMESHEET_BLOCK_SETTINGS_KEY = "global_timesheet_block_settings"
DEFAULT_TIMESHEET_BLOCK_SETTINGS = {
    "first_fortnight_block_day": 14,
    "second_fortnight_block_day": 28,
    "effective_start": "",
    "effective_end": "",
}

ABSENCE_CHARGE_CODES = {
    "adoption_leave": {"code": "955X06", "name": "Adoption Leave"},
    "bereavement_leave": {"code": "955X02", "name": "Bereavement Leave"},
    "casual_leave": {"code": "955X10", "name": "Casual leave"},
    "client_specific_holiday": {"code": "970X01", "name": "Client Specific holiday"},
    "compensatory_off": {"code": "970X01", "name": "Compensatory Off"},
    "contingency_leave": {"code": "955X05", "name": "Contingency Leave"},
    "earned_leave": {"code": "900X00", "name": "Earned Leave"},
    "leave_with_loss_of_pay": {"code": "955X18", "name": "Leave with loss of pay"},
    "maternity_leave": {"code": "955X04", "name": "Maternity Leave"},
    "optional_holiday": {"code": "970X03", "name": "Optional holiday"},
    "other_approved_absence": {"code": "955X00", "name": "Other Approved Absence"},
    "overseas_holiday": {"code": "970X02", "name": "Overseas holiday"},
    "paternity_leave": {"code": "955X08", "name": "Paternity Leave"},
    "public_holiday": {"code": "970X00", "name": "Public holiday"},
    "secondary_caregiver_leave": {"code": "955X19", "name": "Secondary Caregiver Leave"},
    "sick_wellness_leave": {"code": "950X00", "name": "Sick & Wellness Leave"},
    "surrogacy_leave": {"code": "955X07", "name": "Surrogacy Leave"},
}

LEAVE_TYPE_TO_ABSENCE_KEY = {
    "adoption": "adoption_leave",
    "adoption leave": "adoption_leave",
    "bereavement": "bereavement_leave",
    "bereavement leave": "bereavement_leave",
    "casual": "casual_leave",
    "casual leave": "casual_leave",
    "client specific holiday": "client_specific_holiday",
    "compensatory off": "compensatory_off",
    "contingency": "contingency_leave",
    "contingency leave": "contingency_leave",
    "planned": "earned_leave",
    "earned": "earned_leave",
    "earned leave": "earned_leave",
    "lwp": "leave_with_loss_of_pay",
    "lop": "leave_with_loss_of_pay",
    "leave without pay": "leave_with_loss_of_pay",
    "leave with loss of pay": "leave_with_loss_of_pay",
    "maternity": "maternity_leave",
    "maternity leave": "maternity_leave",
    "optional": "optional_holiday",
    "optional holiday": "optional_holiday",
    "other approved absence": "other_approved_absence",
    "overseas holiday": "overseas_holiday",
    "paternity": "paternity_leave",
    "paternity leave": "paternity_leave",
    "secondary caregiver": "secondary_caregiver_leave",
    "secondary caregiver leave": "secondary_caregiver_leave",
    "sick": "sick_wellness_leave",
    "sick leave": "sick_wellness_leave",
    "sick wellness": "sick_wellness_leave",
    "sick and wellness": "sick_wellness_leave",
    "sick & wellness": "sick_wellness_leave",
    "sick & wellness leave": "sick_wellness_leave",
    "surrogacy": "surrogacy_leave",
    "surrogacy leave": "surrogacy_leave",
}

LEAVE_TYPE_DISPLAY_CODES = {
    "casual": "CL",
    "casual leave": "CL",
    "planned": "PL",
    "earned": "PL",
    "earned leave": "PL",
    "sick": "SL",
    "sick leave": "SL",
    "sick wellness": "SL",
    "sick and wellness": "SL",
    "sick & wellness": "SL",
    "sick & wellness leave": "SL",
    "optional": "OH",
    "optional holiday": "OH",
    "lwp": "LWP",
    "lop": "LWP",
    "leave without pay": "LWP",
    "leave with loss of pay": "LWP",
    "early logout": "EL",
    "compensatory off": "CO",
}


# ========================================
# SERIALIZATION HELPERS
# ========================================

def serialize_all(obj):
    """Fully recursive serializer - converts ALL ObjectIds and datetimes."""
    if isinstance(obj, list):
        return [serialize_all(item) for item in obj]
    if isinstance(obj, dict):
        return {k: serialize_all(v) for k, v in obj.items()}
    if isinstance(obj, ObjectId):
        return str(obj)
    if isinstance(obj, datetime):
        return obj.isoformat()
    return obj


def create_notification(
    user_id,
    notification_type,
    message,
    related_timesheet_id=None,
    target=None,
    meta=None,
    reminder_group=None,
    notification_origin="event",
):
    """Create a notification for timesheet actions."""
    try:
        if isinstance(user_id, str):
            user_id = ObjectId(user_id)

        notification = {
            "user_id": user_id,
            "type": notification_type,
            "message": message,
            "read": False,
            "createdAt": datetime.utcnow(),
            "notification_origin": notification_origin,
        }

        if related_timesheet_id:
            if isinstance(related_timesheet_id, str):
                related_timesheet_id = ObjectId(related_timesheet_id)
            notification["related_timesheet_id"] = related_timesheet_id

        if target:
            notification["target"] = target
        if meta:
            notification["meta"] = meta
        if reminder_group:
            notification["reminder_group"] = reminder_group

        mongo.db.notifications.insert_one(notification)
        print(f"✅ Timesheet notification created: {notification_type}")
    except Exception as e:
        print(f"❌ Error creating timesheet notification: {str(e)}")


def build_timesheet_notification_target(timesheet, view="entry"):
    timesheet_id = timesheet if isinstance(timesheet, str) else str((timesheet or {}).get("_id") or (timesheet or {}).get("id") or "")
    target = {
        "section": "timesheets",
        "timesheetId": timesheet_id,
        "activeView": view,
    }
    if isinstance(timesheet, dict):
        target["periodStart"] = timesheet.get("period_start", "")
        target["periodEnd"] = timesheet.get("period_end", "")
    return target


def clear_timesheet_reminder_state(timesheet_id):
    mongo.db.timesheets.update_one(
        {"_id": ObjectId(timesheet_id)},
        {"$unset": {"notification_reminder": ""}},
    )


def set_timesheet_reminder_state(timesheet_id, *, event, mode, recipient_id, target, message, statuses=None):
    recipient = str(recipient_id) if recipient_id else ""
    mongo.db.timesheets.update_one(
        {"_id": ObjectId(timesheet_id)},
        {
            "$set": {
                "notification_reminder": {
                    "event": event,
                    "mode": mode,
                    "recipient_id": recipient,
                    "target": target,
                    "message": message,
                    "statuses": statuses or [],
                    "last_sent_at": datetime.utcnow(),
                    "reminder_group": f"timesheet:{timesheet_id}:{event}:{recipient}",
                }
            }
        },
    )


def send_timesheet_notification_reminders(timesheet_id=None, force=False, reminder_hours=None):
    reminder_hours = int(reminder_hours or os.getenv("TIMESHEET_NOTIFICATION_REMINDER_HOURS", str(TIMESHEET_NOTIFICATION_REMINDER_HOURS)))
    query = {"notification_reminder": {"$exists": True}}
    if timesheet_id:
        query["_id"] = ObjectId(timesheet_id)

    timesheets = list(mongo.db.timesheets.find(query))
    now = datetime.utcnow()
    sent = 0

    for timesheet in timesheets:
        reminder = timesheet.get("notification_reminder") or {}
        recipient_id = reminder.get("recipient_id")
        if not recipient_id:
            continue

        last_sent_at = reminder.get("last_sent_at")
        if not force and last_sent_at and now - last_sent_at < timedelta(hours=reminder_hours):
            continue

        is_active = False
        if reminder.get("mode") == "pending_status":
            is_active = timesheet.get("status") in set(reminder.get("statuses") or [])
        elif reminder.get("mode") == "until_read":
            is_active = mongo.db.notifications.count_documents({
                "user_id": ObjectId(recipient_id),
                "reminder_group": reminder.get("reminder_group"),
                "read": False,
            }) > 0

        if not is_active:
            clear_timesheet_reminder_state(timesheet["_id"])
            continue

        create_notification(
            user_id=recipient_id,
            notification_type=f"timesheet_{reminder.get('event')}_reminder",
            message=reminder.get("message") or "Timesheet reminder",
            related_timesheet_id=timesheet["_id"],
            target=reminder.get("target") or build_timesheet_notification_target(timesheet),
            reminder_group=reminder.get("reminder_group"),
            notification_origin="reminder",
        )
        mongo.db.timesheets.update_one(
            {"_id": timesheet["_id"]},
            {"$set": {"notification_reminder.last_sent_at": now}},
        )
        sent += 1

    return {"tracked": len(timesheets), "notifications_sent": sent, "interval_hours": reminder_hours}


def apply_employee_assignment_snapshot(timesheet_doc, employee):
    """Copy admin-managed assignment fields onto a timesheet document."""
    if not isinstance(timesheet_doc, dict) or not isinstance(employee, dict):
        return timesheet_doc

    timesheet_doc["employee_external_id"] = employee.get("employeeId", "") or ""
    timesheet_doc["employee_work_location"] = employee.get("workLocation", "") or ""
    timesheet_doc["employee_assigned_location"] = (
        employee.get("assignedLocation")
        or employee.get("costCenter")
        or employee.get("workLocation")
        or ""
    )
    timesheet_doc["employee_company_code"] = employee.get("companyCode", "") or ""
    timesheet_doc["employee_cost_center"] = employee.get("costCenter", "") or ""
    return timesheet_doc


def enrich_timesheet_with_employee_assignments(timesheet_doc):
    """Backfill assignment metadata for older timesheets when reading."""
    if not isinstance(timesheet_doc, dict):
        return timesheet_doc

    employee_id = timesheet_doc.get("employee_id")
    if not employee_id:
        return timesheet_doc

    if (
        timesheet_doc.get("employee_work_location")
        and timesheet_doc.get("employee_assigned_location")
        and timesheet_doc.get("employee_company_code")
        and timesheet_doc.get("employee_external_id")
    ):
        return timesheet_doc

    try:
        employee_lookup_id = employee_id if isinstance(employee_id, ObjectId) else ObjectId(employee_id)
        employee = mongo.db.users.find_one({"_id": employee_lookup_id})
        if employee:
            apply_employee_assignment_snapshot(timesheet_doc, employee)
    except Exception:
        pass

    return timesheet_doc


def validate_daily_work_hours(entries):
    """Block more than 9 submitted work hours on any single day."""
    daily_totals = {}
    for entry in entries or []:
        if entry.get("entry_type", "work") != "work":
            continue
        date_key = entry.get("date")
        if not date_key:
            continue
        try:
            hours = float(entry.get("hours") or 0)
        except (TypeError, ValueError):
            return f"Invalid hours on {date_key}"
        if hours < 0:
            return f"Hours cannot be negative on {date_key}"
        if hours > 9:
            return f"Working hours for any charge code cannot exceed 9 hours on {date_key}"
        daily_totals[date_key] = daily_totals.get(date_key, 0) + hours

    over_limit = [
        f"{date_key} ({total:g}h)"
        for date_key, total in sorted(daily_totals.items())
        if total > 9
    ]
    if over_limit:
        return "Total working hours across all charge codes cannot exceed 9 hours per day: " + ", ".join(over_limit)
    return None


def normalize_date_key(value):
    """Normalize stored datetime/string values to YYYY-MM-DD for comparisons."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value)[:10]


def normalize_email_address(value):
    return str(value or "").strip().lower()


def normalize_email_list(value):
    if not value:
        return []
    if isinstance(value, str):
        raw_items = value.replace("\n", ",").split(",")
    elif isinstance(value, list):
        raw_items = value
    else:
        raw_items = [value]

    normalized = []
    seen = set()
    for item in raw_items:
        email = normalize_email_address(item)
        if not email or "@" not in email or email in seen:
            continue
        seen.add(email)
        normalized.append(email)
    return normalized


def find_user_by_email(email):
    normalized_email = normalize_email_address(email)
    if not normalized_email:
        return None
    return mongo.db.users.find_one({"email": {"$regex": f"^{normalized_email}$", "$options": "i"}})


def build_timesheet_preference_document(employee_id, period_start, period_end, payload, managed_by=None):
    now = datetime.utcnow()
    return {
        "employee_id": employee_id,
        "period_start": period_start,
        "period_end": period_end,
        "reviewers": normalize_email_list(payload.get("reviewers")),
        "notifications": normalize_email_list(payload.get("notifications")),
        "delegates": normalize_email_list(payload.get("delegates")),
        "approvers": normalize_email_list(payload.get("approvers")),
        "managed_by_admin_id": managed_by.get("_id") if managed_by else None,
        "managed_by_admin_name": (managed_by or {}).get("name", ""),
        "updated_at": now,
    }


def get_timesheet_preference(employee_id, period_start, period_end):
    if not employee_id or not period_start or not period_end:
        return None
    return mongo.db.timesheet_preferences.find_one({
        "employee_id": employee_id,
        "period_start": period_start,
        "period_end": period_end,
    })


def apply_timesheet_workflow_preference(timesheet_doc, employee=None):
    if not isinstance(timesheet_doc, dict):
        return timesheet_doc

    preference = get_timesheet_preference(
        timesheet_doc.get("employee_id"),
        timesheet_doc.get("period_start"),
        timesheet_doc.get("period_end"),
    )
    if not preference:
        return timesheet_doc

    approver_emails = normalize_email_list(preference.get("approvers"))
    notification_emails = normalize_email_list(preference.get("notifications"))
    reviewer_emails = normalize_email_list(preference.get("reviewers"))
    delegate_emails = normalize_email_list(preference.get("delegates"))

    matched_approver = None
    for approver_email in approver_emails:
        matched_approver = find_user_by_email(approver_email)
        if matched_approver:
            break

    if matched_approver:
        timesheet_doc["reporting_lead_id"] = matched_approver["_id"]
        timesheet_doc["reporting_lead_name"] = matched_approver.get("name", "")
        timesheet_doc["reporting_lead_email"] = matched_approver.get("email", "")
    elif employee and employee.get("reportsTo"):
        fallback_lead = mongo.db.users.find_one({"_id": employee.get("reportsTo")})
        if fallback_lead:
            timesheet_doc["reporting_lead_name"] = fallback_lead.get("name", "")
            timesheet_doc["reporting_lead_email"] = fallback_lead.get("email", "")

    timesheet_doc["timesheet_workflow_preferences"] = {
        "reviewers": reviewer_emails,
        "notifications": notification_emails,
        "delegates": delegate_emails,
        "approvers": approver_emails,
        "managed_by_admin_id": preference.get("managed_by_admin_id"),
        "managed_by_admin_name": preference.get("managed_by_admin_name", ""),
        "updated_at": preference.get("updated_at"),
    }
    return timesheet_doc


def queue_timesheet_submission_emails(timesheet_doc, employee=None):
    preference = (timesheet_doc or {}).get("timesheet_workflow_preferences") or {}
    approvers = normalize_email_list(preference.get("approvers"))
    notifications = normalize_email_list(preference.get("notifications"))
    reviewers = normalize_email_list(preference.get("reviewers"))
    delegates = normalize_email_list(preference.get("delegates"))

    if not approvers and not notifications:
        return 0

    employee = employee or mongo.db.users.find_one({"_id": timesheet_doc.get("employee_id")})
    tenant_id = resolve_tenant_id(timesheet_doc, employee)
    context = {
        "employee_name": timesheet_doc.get("employee_name", "Employee"),
        "employee_email": timesheet_doc.get("employee_email", ""),
        "period_start": timesheet_doc.get("period_start", ""),
        "period_end": timesheet_doc.get("period_end", ""),
        "total_hours": timesheet_doc.get("total_hours", 0),
        "work_hours": timesheet_doc.get("work_hours", 0),
        "reviewers": reviewers,
        "notifications": notifications,
        "delegates": delegates,
        "approvers": approvers,
        "managed_by_admin_name": preference.get("managed_by_admin_name", ""),
    }

    queued = 0
    timesheet_id = timesheet_doc.get("_id")

    if approvers:
        enqueue_mail(
            tenant_id=tenant_id,
            employee_id=timesheet_doc.get("employee_id"),
            leave_id=str(timesheet_id) if timesheet_id else None,
            mail_type="timesheet_approval_request",
            recipients=approvers,
            cc=reviewers + delegates,
            subject=f"Timesheet approval request: {context['employee_name']}",
            template_name="timesheet_submitted.html",
            context={**context, "audience": "approver"},
            idempotency_key=f"{tenant_id}:timesheet:{timesheet_id}:submit:approvers:v1",
            metadata={"timesheet_id": str(timesheet_id) if timesheet_id else ""},
        )
        queued += 1

    if notifications:
        enqueue_mail(
            tenant_id=tenant_id,
            employee_id=timesheet_doc.get("employee_id"),
            leave_id=str(timesheet_id) if timesheet_id else None,
            mail_type="timesheet_submission_notification",
            recipients=notifications,
            cc=delegates,
            subject=f"Timesheet submitted: {context['employee_name']}",
            template_name="timesheet_submitted.html",
            context={**context, "audience": "notification"},
            idempotency_key=f"{tenant_id}:timesheet:{timesheet_id}:submit:notifications:v1",
            metadata={"timesheet_id": str(timesheet_id) if timesheet_id else ""},
        )
        queued += 1

    return queued


def get_timesheet_manager_hierarchy(employee_id):
    """Return the reporting chain for timesheet escalation and multi-step approval."""
    hierarchy = []
    current_id = employee_id
    visited = set()

    while current_id:
        current_key = str(current_id)
        if current_key in visited:
            break
        visited.add(current_key)

        employee = mongo.db.users.find_one({"_id": ObjectId(current_id)})
        if not employee or not employee.get("reportsTo"):
            break

        manager_id = employee.get("reportsTo")
        manager = mongo.db.users.find_one({"_id": manager_id})
        if not manager:
            break

        hierarchy.append({
            "_id": manager["_id"],
            "name": manager.get("name", "Unknown"),
            "email": manager.get("email", ""),
            "role": manager.get("role", "Manager"),
        })
        current_id = manager_id

    return hierarchy


def get_timesheet_approval_chain(timesheet, employee=None):
    """Resolve the ordered approver chain for a timesheet."""
    if not isinstance(timesheet, dict):
        return []

    employee = employee or mongo.db.users.find_one({"_id": timesheet.get("employee_id")})
    chain = []
    seen = set()

    def append_user(user_doc):
        if not user_doc or not user_doc.get("_id"):
            return
        key = str(user_doc["_id"])
        if key in seen:
            return
        seen.add(key)
        chain.append({
            "_id": user_doc["_id"],
            "name": user_doc.get("name", ""),
            "email": user_doc.get("email", ""),
            "role": user_doc.get("role", ""),
        })

    workflow = (timesheet.get("timesheet_workflow_preferences") or {})
    for approver_email in normalize_email_list(workflow.get("approvers")):
        append_user(find_user_by_email(approver_email))

    reporting_lead_id = timesheet.get("reporting_lead_id") or (employee or {}).get("reportsTo")
    if reporting_lead_id:
        append_user(mongo.db.users.find_one({"_id": reporting_lead_id}))

    manager_id = timesheet.get("manager_id") or (employee or {}).get("peopleLead")
    if manager_id:
        append_user(mongo.db.users.find_one({"_id": manager_id}))

    for manager in get_timesheet_manager_hierarchy(timesheet.get("employee_id")):
        append_user(manager)

    return chain


def get_timesheet_current_approver(timesheet, employee=None):
    chain = get_timesheet_approval_chain(timesheet, employee=employee)
    current_id = str((timesheet or {}).get("current_approver_id") or "")
    if current_id:
        for approver in chain:
            if str(approver.get("_id")) == current_id:
                return approver
    return chain[0] if chain else None


def get_next_timesheet_approver(timesheet, employee=None):
    chain = get_timesheet_approval_chain(timesheet, employee=employee)
    if not chain:
        return None

    current_id = str((timesheet or {}).get("current_approver_id") or "")
    if not current_id:
        return chain[0]

    for index, approver in enumerate(chain):
        if str(approver.get("_id")) == current_id:
            return chain[index + 1] if index + 1 < len(chain) else None
    return None


def notify_timesheet_pending_approver(timesheet, approver_id, *, event, message, notification_type="timesheet_submitted"):
    if not approver_id:
        return
    target = build_timesheet_notification_target(timesheet, view="approvals")
    create_notification(
        user_id=approver_id,
        notification_type=notification_type,
        message=message,
        related_timesheet_id=timesheet.get("_id"),
        target=target,
        reminder_group=f"timesheet:{timesheet.get('_id')}:{event}:{approver_id}",
    )
    set_timesheet_reminder_state(
        timesheet.get("_id"),
        event=event,
        mode="pending_status",
        recipient_id=approver_id,
        target=target,
        message=message,
        statuses=["pending_lead", "pending_manager"],
    )


def finalize_timesheet_approval(timesheet, approver_id, approver_name, approval_entry, *, final_stage_label="approver"):
    now = datetime.utcnow()
    mongo.db.timesheets.update_one(
        {"_id": timesheet["_id"]},
        {
            "$set": {
                "status": "approved",
                "current_approver_id": None,
                "lead_approved_at": timesheet.get("lead_approved_at") or now,
                "lead_approved_by": timesheet.get("lead_approved_by") or approver_name,
                "lead_approved_by_id": timesheet.get("lead_approved_by_id") or approver_id,
                "manager_approved_at": now if approval_entry.get("stage") == "manager" else timesheet.get("manager_approved_at"),
                "manager_approved_by": approver_name if approval_entry.get("stage") == "manager" else timesheet.get("manager_approved_by"),
                "manager_approved_by_id": approver_id if approval_entry.get("stage") == "manager" else timesheet.get("manager_approved_by_id"),
                "is_locked": True,
                "requires_reapproval": False,
                "updated_at": now,
            },
            "$push": {"approval_history": approval_entry},
        },
    )

    approval_events = [
        item for item in (timesheet.get("approval_history") or [])
        if item.get("action") == "approved"
    ]
    approval_event = "reapproved" if approval_events else "approved"
    employee_target = build_timesheet_notification_target(timesheet, view="entry")
    approval_message = (
        f"Your timesheet ({timesheet.get('period_start')} to {timesheet.get('period_end')}) "
        f"has been approved by your {final_stage_label} and is now locked"
    )
    create_notification(
        user_id=timesheet["employee_id"],
        notification_type="timesheet_approved",
        message=approval_message,
        related_timesheet_id=timesheet["_id"],
        target=employee_target,
        reminder_group=f"timesheet:{timesheet['_id']}:{approval_event}:{timesheet['employee_id']}",
    )
    set_timesheet_reminder_state(
        timesheet["_id"],
        event=approval_event,
        mode="until_read",
        recipient_id=timesheet["employee_id"],
        target=employee_target,
        message=approval_message,
    )


def move_timesheet_to_next_approver(timesheet, next_approver, approval_entry, *, escalation=False):
    now = datetime.utcnow()
    next_status = "pending_manager"
    update_data = {
        "status": next_status,
        "current_approver_id": next_approver["_id"],
        "current_approver_name": next_approver.get("name", ""),
        "current_approver_email": next_approver.get("email", ""),
        "updated_at": now,
    }
    if escalation:
        update_data["escalation_level"] = int(timesheet.get("escalation_level", 0) or 0) + 1
        update_data["escalated_on"] = now
        update_data["previous_approver_id"] = timesheet.get("current_approver_id")

    push_data = {"approval_history": approval_entry}
    if escalation:
        push_data["escalation_history"] = {
            "from_approver_id": timesheet.get("current_approver_id"),
            "to_approver_id": next_approver["_id"],
            "to_approver_name": next_approver.get("name", ""),
            "escalated_at": now,
            "level": int(timesheet.get("escalation_level", 0) or 0) + 1,
        }

    mongo.db.timesheets.update_one(
        {"_id": timesheet["_id"]},
        {"$set": update_data, "$push": push_data},
    )

    action_label = "escalated" if escalation else "forwarded"
    message = (
        f"{timesheet.get('employee_name', 'An employee')} submitted a timesheet for "
        f"{timesheet.get('period_start')} to {timesheet.get('period_end')} and it was {action_label} to you for approval"
    )
    notify_timesheet_pending_approver(
        {**timesheet, "_id": timesheet["_id"], "status": next_status},
        next_approver["_id"],
        event="escalated" if escalation else "submitted",
        message=message,
        notification_type="timesheet_escalated" if escalation else "timesheet_submitted",
    )


def escalate_timesheet_request(timesheet_id):
    try:
        timesheet = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not timesheet or timesheet.get("status") not in ("pending_lead", "pending_manager"):
            return False

        employee = mongo.db.users.find_one({"_id": timesheet.get("employee_id")})
        next_approver = get_next_timesheet_approver(timesheet, employee=employee)
        current_approver = get_timesheet_current_approver(timesheet, employee=employee)

        if not next_approver:
            admins = list(mongo.db.users.find({"role": "Admin"}).sort("name", 1))
            current_is_admin = bool(current_approver and current_approver.get("role") == "Admin")
            if current_is_admin or not admins:
                return False
            next_approver = {
                "_id": admins[0]["_id"],
                "name": admins[0].get("name", "Admin"),
                "email": admins[0].get("email", ""),
                "role": "Admin",
            }

        approval_entry = {
            "stage": "manager" if timesheet.get("status") == "pending_manager" else "lead",
            "action": "escalated",
            "approver_id": current_approver.get("_id") if current_approver else None,
            "approver_name": current_approver.get("name", "") if current_approver else "System",
            "comments": f"Auto-escalated to {next_approver.get('name', 'approver')}",
            "timestamp": datetime.utcnow(),
        }
        move_timesheet_to_next_approver(timesheet, next_approver, approval_entry, escalation=True)
        return True
    except Exception as e:
        print(f"❌ Error escalating timesheet {timesheet_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def check_timesheet_escalations():
    try:
        current_time = now_ist().replace(tzinfo=None)
        pending_timesheets = list(mongo.db.timesheets.find({"status": {"$in": ["pending_lead", "pending_manager"]}}))
        escalated_count = 0

        for timesheet in pending_timesheets:
            escalation_level = int(timesheet.get("escalation_level", 0) or 0)
            reference_date = timesheet.get("escalated_on") or timesheet.get("submitted_at")
            if not reference_date:
                continue

            if getattr(reference_date, "tzinfo", None) is not None:
                reference_date = reference_date.astimezone(now_ist().tzinfo).replace(tzinfo=None)

            days_pending = (current_time - reference_date).days
            should_escalate = False
            if escalation_level == 0 and days_pending >= 2:
                should_escalate = True
            elif escalation_level > 0 and days_pending >= 1:
                should_escalate = True

            if should_escalate and escalate_timesheet_request(str(timesheet["_id"])):
                escalated_count += 1

        return {
            "message": "Timesheet escalation check completed",
            "total_pending": len(pending_timesheets),
            "escalated_count": escalated_count,
        }
    except Exception as e:
        print(f"❌ Error checking timesheet escalations: {str(e)}")
        raise


@timesheet_bp.route("/check_escalations", methods=["POST"])
def run_timesheet_escalation_check():
    try:
        result = check_timesheet_escalations()
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def get_timesheet_period_bounds(timesheet):
    """Parse the stored period bounds for fortnight-aware edit rules."""
    try:
        start_key = normalize_date_key(timesheet.get("period_start"))
        end_key = normalize_date_key(timesheet.get("period_end"))
        if not start_key or not end_key:
            return None, None
        return (
            datetime.strptime(start_key, "%Y-%m-%d").date(),
            datetime.strptime(end_key, "%Y-%m-%d").date(),
        )
    except Exception:
        return None, None


def get_period_bounds(period_start, period_end):
    try:
        start_key = normalize_date_key(period_start)
        end_key = normalize_date_key(period_end)
        if not start_key or not end_key:
            return None, None
        return (
            datetime.strptime(start_key, "%Y-%m-%d").date(),
            datetime.strptime(end_key, "%Y-%m-%d").date(),
        )
    except Exception:
        return None, None


def clamp_day_for_month(source_date, day):
    try:
        day = int(day)
        next_month = (source_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_end = next_month - timedelta(days=1)
        return source_date.replace(day=max(1, min(day, month_end.day)))
    except Exception:
        return None


def normalize_timesheet_block_settings(settings=None):
    settings = settings or {}
    normalized = dict(DEFAULT_TIMESHEET_BLOCK_SETTINGS)
    for key, min_day, max_day in (
        ("first_fortnight_block_day", 1, 31),
        ("second_fortnight_block_day", 1, 31),
    ):
        try:
            value = int(settings.get(key, normalized[key]))
        except Exception:
            value = normalized[key]
        normalized[key] = max(min_day, min(value, max_day))
    for key in ("effective_start", "effective_end"):
        normalized[key] = normalize_date_key(settings.get(key)) or ""
    if normalized["effective_start"] and normalized["effective_end"] and normalized["effective_start"] > normalized["effective_end"]:
        normalized["effective_start"], normalized["effective_end"] = normalized["effective_end"], normalized["effective_start"]
    return normalized


def format_ordinal(day):
    day = int(day)
    if 10 <= day % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return f"{day}{suffix}"


def get_timesheet_block_settings():
    settings = mongo.db.system_settings.find_one({"key": TIMESHEET_BLOCK_SETTINGS_KEY}) or {}
    return normalize_timesheet_block_settings(settings.get("value") or settings)


def rule_matches_period(settings, period_start_date, period_end_date):
    if not period_start_date or not period_end_date:
        return False

    effective_start = settings.get("effective_start")
    effective_end = settings.get("effective_end")
    start_bound = datetime.strptime(effective_start, "%Y-%m-%d").date() if effective_start else None
    end_bound = datetime.strptime(effective_end, "%Y-%m-%d").date() if effective_end else None

    if start_bound and period_end_date < start_bound:
        return False
    if end_bound and period_start_date > end_bound:
        return False
    return True


def serialize_timesheet_block_override(override):
    if not override:
        return None
    return {
        "_id": str(override.get("_id")),
        "employee_id": str(override.get("employee_id")),
        "employee_name": override.get("employee_name", ""),
        "employee_email": override.get("employee_email", ""),
        "first_fortnight_block_day": override.get("first_fortnight_block_day", 14),
        "second_fortnight_block_day": override.get("second_fortnight_block_day", 28),
        "effective_start": override.get("effective_start", ""),
        "effective_end": override.get("effective_end", ""),
        "is_active": override.get("is_active", True),
        "notes": override.get("notes", ""),
        "updated_at": override.get("updated_at").isoformat() if isinstance(override.get("updated_at"), datetime) else override.get("updated_at"),
    }


def serialize_timesheet_block_history(item):
    if not item:
        return None
    return {
        "_id": str(item.get("_id")),
        "action": item.get("action", ""),
        "scope": item.get("scope", ""),
        "employee_id": str(item.get("employee_id")) if item.get("employee_id") else "",
        "employee_name": item.get("employee_name", ""),
        "employee_email": item.get("employee_email", ""),
        "first_fortnight_block_day": item.get("first_fortnight_block_day", 14),
        "second_fortnight_block_day": item.get("second_fortnight_block_day", 28),
        "effective_start": item.get("effective_start", ""),
        "effective_end": item.get("effective_end", ""),
        "notes": item.get("notes", ""),
        "changed_by_name": item.get("changed_by_name", ""),
        "changed_by_email": item.get("changed_by_email", ""),
        "changed_at": item.get("changed_at").isoformat() if isinstance(item.get("changed_at"), datetime) else item.get("changed_at"),
    }


def record_timesheet_block_history(action, scope, settings, requester, employee=None, notes=""):
    history_doc = {
        "action": action,
        "scope": scope,
        **normalize_timesheet_block_settings(settings),
        "notes": notes or "",
        "changed_at": datetime.utcnow(),
        "changed_by": requester.get("_id") if requester else None,
        "changed_by_name": (requester or {}).get("name", ""),
        "changed_by_email": (requester or {}).get("email", ""),
    }
    if employee:
        history_doc.update({
            "employee_id": employee.get("_id"),
            "employee_name": employee.get("name", ""),
            "employee_email": employee.get("email", ""),
        })
    mongo.db.timesheet_block_history.insert_one(history_doc)


def get_matching_employee_block_override(employee_id, period_start_date, period_end_date):
    if not employee_id:
        return None
    try:
        employee_obj_id = employee_id if isinstance(employee_id, ObjectId) else ObjectId(employee_id)
    except Exception:
        return None

    overrides = list(mongo.db.timesheet_block_overrides.find({
        "employee_id": employee_obj_id,
        "is_active": True,
    }).sort("updated_at", -1))
    for override in overrides:
        settings = normalize_timesheet_block_settings(override)
        if rule_matches_period(settings, period_start_date, period_end_date):
            return settings, override
    return None


def resolve_timesheet_block_settings(employee_id, period_start, period_end):
    start_date, end_date = get_period_bounds(period_start, period_end)
    if not start_date or not end_date:
        return None, None

    employee_match = get_matching_employee_block_override(employee_id, start_date, end_date)
    if employee_match:
        return employee_match

    global_settings = get_timesheet_block_settings()
    if not rule_matches_period(global_settings, start_date, end_date):
        return None, None
    return global_settings, None


def get_fortnight_deadline_date(period_start, period_end, employee_id=None):
    start_date, end_date = get_period_bounds(period_start, period_end)
    if not start_date or not end_date:
        return None

    settings, _ = resolve_timesheet_block_settings(employee_id, period_start, period_end)
    if not settings:
        return None
    if start_date.day == 1 and end_date.day == 15:
        return clamp_day_for_month(start_date, settings["first_fortnight_block_day"])
    if start_date.day == 16:
        return clamp_day_for_month(start_date, settings["second_fortnight_block_day"])
    return None


def get_fortnight_deadline_label(period_start, period_end, employee_id=None):
    deadline = get_fortnight_deadline_date(period_start, period_end, employee_id=employee_id)
    start_date, end_date = get_period_bounds(period_start, period_end)
    if not deadline or not start_date or not end_date:
        return ""

    if start_date.day == 1 and end_date.day == 15:
        return f"This first-fortnight timesheet is blocked from the {format_ordinal(deadline.day)} onward unless an admin unblocks it."
    if start_date.day == 16:
        return f"This second-fortnight timesheet is blocked from the {format_ordinal(deadline.day)} onward unless an admin unblocks it."
    return ""


def get_timesheet_period_unlock(employee_id, period_start, period_end):
    if not employee_id or not period_start or not period_end:
        return None
    return mongo.db.timesheet_period_unlocks.find_one({
        "employee_id": employee_id,
        "period_start": period_start,
        "period_end": period_end,
        "is_active": True,
    })


def is_timesheet_period_unlocked(employee_id, period_start, period_end):
    return bool(get_timesheet_period_unlock(employee_id, period_start, period_end))


def is_fortnight_entry_blocked(employee_id, period_start, period_end, reference=None):
    deadline = get_fortnight_deadline_date(period_start, period_end, employee_id=employee_id)
    if not deadline:
        return False

    current_date = (reference or now_ist()).date()
    if current_date < deadline:
        return False

    return not is_timesheet_period_unlocked(employee_id, period_start, period_end)


def get_period_access_state(employee_id, period_start, period_end, reference=None):
    unlock = get_timesheet_period_unlock(employee_id, period_start, period_end)
    deadline = get_fortnight_deadline_date(period_start, period_end, employee_id=employee_id)
    settings, override_record = resolve_timesheet_block_settings(employee_id, period_start, period_end)
    blocked = is_fortnight_entry_blocked(employee_id, period_start, period_end, reference=reference)
    return {
        "employee_id": employee_id,
        "period_start": period_start,
        "period_end": period_end,
        "entry_deadline_date": deadline.strftime("%Y-%m-%d") if deadline else "",
        "entry_deadline_label": get_fortnight_deadline_label(period_start, period_end, employee_id=employee_id),
        "entry_blocked": blocked,
        "unlock_active": bool(unlock),
        "unlock_record": unlock,
        "block_rule_scope": "employee" if override_record else ("global" if settings else "none"),
        "block_rule": settings or {},
        "block_rule_override": serialize_timesheet_block_override(override_record) if override_record else None,
    }


def is_approved_edit_window_open(timesheet, reference=None):
    """Allow approved employee edits only on the configured payroll correction days."""
    if not timesheet or timesheet.get("status") != "approved":
        return False

    return is_correction_window_open_for_timesheet(timesheet, reference=reference)


def is_correction_window_open_for_timesheet(timesheet, reference=None):
    """Return True when the fortnight's correction dates are open."""
    if not timesheet:
        return False

    period_start, period_end = get_timesheet_period_bounds(timesheet)
    if not period_start or not period_end:
        return False

    current_date = (reference or now_ist()).date()
    if current_date.year != period_start.year or current_date.month != period_start.month:
        return False

    if period_start.day == 1 and period_end.day == 15:
        return current_date.day in (13, 14)
    if period_start.day == 16 and period_end.day >= 28:
        return current_date.day in (26, 27)

    return False


def get_approved_edit_window_label(timesheet):
    period_start, period_end = get_timesheet_period_bounds(timesheet)
    if not period_start or not period_end:
        return ""
    if period_start.day == 1 and period_end.day == 15:
        return "Editable on the 13th and 14th of the same month after approval."
    if period_start.day == 16:
        return "Editable on the 26th and 27th of the same month after approval."
    return ""


def is_employee_editable_timesheet(timesheet):
    period_blocked = is_fortnight_entry_blocked(
        timesheet.get("employee_id"),
        timesheet.get("period_start"),
        timesheet.get("period_end"),
    )
    status = (timesheet or {}).get("status")
    if status in ("rejected_by_lead", "rejected_by_manager"):
        return not period_blocked
    if status == "draft":
        if timesheet.get("reopened_from_approved"):
            return is_correction_window_open_for_timesheet(timesheet) or is_timesheet_period_unlocked(
                timesheet.get("employee_id"),
                timesheet.get("period_start"),
                timesheet.get("period_end"),
            )
        return not period_blocked
    if status == "approved":
        return is_correction_window_open_for_timesheet(timesheet) or is_timesheet_period_unlocked(
            timesheet.get("employee_id"),
            timesheet.get("period_start"),
            timesheet.get("period_end"),
        )
    return False


def build_resubmission_update(ts, update_data):
    """Convert an approved timesheet into a draft revision during the reopen window."""
    if ts.get("status") != "approved":
        return update_data

    next_data = dict(update_data)
    next_data.update({
        "status": "draft",
        "is_locked": False,
        "reopened_from_approved": True,
        "requires_reapproval": True,
        "approved_edit_window_used_at": datetime.utcnow(),
    })
    return next_data


def annotate_timesheet_editability(timesheet):
    if not isinstance(timesheet, dict):
        return timesheet

    current_approver_id = timesheet.get("current_approver_id")
    if current_approver_id and not timesheet.get("current_approver_name"):
        current_approver = mongo.db.users.find_one({"_id": current_approver_id})
        if current_approver:
            timesheet["current_approver_name"] = current_approver.get("name", "")
            timesheet["current_approver_email"] = current_approver.get("email", "")

    access_state = get_period_access_state(
        timesheet.get("employee_id"),
        timesheet.get("period_start"),
        timesheet.get("period_end"),
    )
    timesheet["approved_edit_window_open"] = is_approved_edit_window_open(timesheet)
    timesheet["approved_edit_window_label"] = get_approved_edit_window_label(timesheet)
    timesheet["is_employee_editable"] = is_employee_editable_timesheet(timesheet)
    timesheet["period_entry_blocked"] = access_state["entry_blocked"]
    timesheet["period_entry_deadline_date"] = access_state["entry_deadline_date"]
    timesheet["period_entry_deadline_label"] = access_state["entry_deadline_label"]
    timesheet["period_unlock_active"] = access_state["unlock_active"]
    if access_state["unlock_record"]:
        timesheet["period_unlock_record"] = access_state["unlock_record"]
    return timesheet


def is_weekday_date(date_key):
    try:
        return datetime.strptime(date_key, "%Y-%m-%d").weekday() < 5
    except Exception:
        return False


def daterange_keys(start_key, end_key):
    start = datetime.strptime(start_key, "%Y-%m-%d")
    end = datetime.strptime(end_key, "%Y-%m-%d")
    current = start
    while current <= end:
        yield current.strftime("%Y-%m-%d")
        current += timedelta(days=1)


def normalize_daily_adjustments(adjustments, period_start, period_end, field_label):
    """Validate editable daily adjustment rows such as overtime and holiday payout."""
    if adjustments in (None, ""):
        return {}, None
    if not isinstance(adjustments, dict):
        return {}, f"{field_label} must be keyed by date"

    valid_dates = set(daterange_keys(period_start, period_end))
    normalized = {}
    for raw_date, raw_hours in adjustments.items():
        date_key = normalize_date_key(raw_date)
        if not date_key or date_key not in valid_dates:
            return {}, f"{field_label} contains a date outside the timesheet period"
        if raw_hours in (None, ""):
            continue

        try:
            hours = float(raw_hours or 0)
        except (TypeError, ValueError):
            return {}, f"{field_label} has invalid hours on {date_key}"

        if hours < 0:
            return {}, f"{field_label} cannot be negative on {date_key}"
        if hours > WORKDAY_HOURS:
            return {}, f"{field_label} cannot exceed {WORKDAY_HOURS:g} hours on {date_key}"

        normalized[date_key] = round(hours, 2)

    return normalized, None


def normalize_work_schedule(adjustments, period_start, period_end):
    """Validate editable daily work schedule values stored with a timesheet."""
    return normalize_daily_adjustments(adjustments, period_start, period_end, "Work schedule")


def normalize_location_map(locations, period_start, period_end):
    """Normalize optional date-keyed location metadata for a timesheet period."""
    if not isinstance(locations, dict):
        return {}

    valid_dates = set(daterange_keys(period_start, period_end))
    normalized = {}
    for raw_date, raw_location in locations.items():
        date_key = normalize_date_key(raw_date)
        if date_key not in valid_dates:
            continue
        location = str(raw_location or "").strip()
        if location:
            normalized[date_key] = location
    return normalized


def normalize_absence_label(value):
    return " ".join(
        str(value or "")
        .replace("&", " and ")
        .replace("-", " ")
        .replace("_", " ")
        .strip()
        .lower()
        .split()
    )


def is_early_logout_leave_type(value):
    return normalize_absence_label(value) == "early logout"


def is_lop_entry(entry):
    """Identify loss-of-pay rows generated from approved leave/timesheet entries."""
    values = [
        entry.get("leave_type"),
        entry.get("description"),
        entry.get("charge_code_name"),
        entry.get("display_code"),
        entry.get("code"),
        entry.get("charge_code"),
    ]
    normalized_values = {normalize_absence_label(value) for value in values if value not in (None, "")}
    return bool(
        {"lwp", "lop", "leave without pay", "leave with loss of pay"} & normalized_values
        or "955x18" in {str(value or "").strip().lower() for value in values}
    )


def parse_report_date(value, fallback=None):
    if not value:
        return fallback
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d")
    except Exception:
        return fallback


def report_date_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value or "")[:10]


def count_weekdays_between(start_key, end_key):
    if not start_key or not end_key:
        return 0
    return sum(1 for item in daterange_keys(start_key, end_key) if is_weekday_date(item))


def get_employee_status_label(employee):
    if employee.get("isActive") is False or employee.get("active") is False:
        return "Inactive"
    return employee.get("status") or "Active"


def build_lop_report_rows(start_key, end_key, filters=None):
    filters = filters or {}
    employee_query = {"role": {"$ne": "Admin"}}
    department = (filters.get("department") or "").strip()
    employee_status = (filters.get("employee_status") or "all").strip().lower()
    search = (filters.get("search") or "").strip().lower()

    if department and department != "all":
        employee_query["department"] = department

    employees = list(mongo.db.users.find(employee_query).sort("name", 1))
    if employee_status == "active":
        employees = [emp for emp in employees if get_employee_status_label(emp).lower() == "active"]
    elif employee_status == "inactive":
        employees = [emp for emp in employees if get_employee_status_label(emp).lower() != "active"]

    if search:
        employees = [
            emp for emp in employees
            if search in " ".join([
                str(emp.get("employeeId", "")),
                str(emp.get("name", "")),
                str(emp.get("email", "")),
                str(emp.get("department", "")),
            ]).lower()
        ]

    employee_ids = [emp["_id"] for emp in employees]
    timesheets_by_employee = {str(emp_id): [] for emp_id in employee_ids}
    if employee_ids:
        timesheets = list(mongo.db.timesheets.find({
            "employee_id": {"$in": employee_ids},
            "period_start": {"$lte": end_key},
            "period_end": {"$gte": start_key},
        }))
        for ts in timesheets:
            refreshed = refresh_timesheet_for_read(ts)
            timesheets_by_employee.setdefault(str(refreshed.get("employee_id")), []).append(refreshed)

    rows = []
    total_lop_days = 0.0
    total_actual_lop_days = 0.0
    total_adjusted_lop_days = 0.0
    for employee in employees:
        lop_hours = 0.0
        for timesheet in timesheets_by_employee.get(str(employee["_id"]), []):
            for entry in timesheet.get("entries", []) or []:
                entry_date = report_date_key(entry.get("date"))
                if not entry_date or entry_date < start_key or entry_date > end_key:
                    continue
                if not is_lop_entry(entry):
                    continue
                try:
                    lop_hours += float(entry.get("hours") or 0)
                except (TypeError, ValueError):
                    continue

        actual_lop_days = round(lop_hours / WORKDAY_HOURS, 2) if WORKDAY_HOURS else 0
        lop_days = round(actual_lop_days, 2)
        adjusted_lop_days = 0.0
        total_lop_days += lop_days
        total_actual_lop_days += actual_lop_days
        total_adjusted_lop_days += adjusted_lop_days
        rows.append({
            "employee_id": employee.get("employeeId", "") or str(employee["_id"]),
            "employee_name": employee.get("name", ""),
            "employee_email": employee.get("email", ""),
            "employee_status": get_employee_status_label(employee),
            "department": employee.get("department", ""),
            "working_days": count_weekdays_between(start_key, end_key),
            "adjustment_days": adjusted_lop_days,
            "lop_days": lop_days,
            "actual_lop_days": actual_lop_days,
        })

    return {
        "rows": rows,
        "totals": {
            "employees": len(rows),
            "lop_days": round(total_lop_days, 2),
            "adjusted_lop_days": round(total_adjusted_lop_days, 2),
            "actual_lop_days": round(total_actual_lop_days, 2),
        },
    }


def create_lop_report_workbook(start_key, end_key, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "lop_summary"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Naxrita Solutions Private Limited\nLoss Of Pay Summary\n\nFrom {datetime.strptime(start_key, '%Y-%m-%d').strftime('%d/%m/%Y')} To {datetime.strptime(end_key, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 68

    headers = [
        "Employee ID",
        "Employee Name",
        "Employee Email ID",
        "Employee Status",
        "Working Days",
        "LOP Days",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F7F8FC")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([
            row["employee_id"],
            row["employee_name"],
            row["employee_email"],
            row["employee_status"],
            row["working_days"],
            row["lop_days"],
        ])

    widths = [18, 28, 34, 18, 16, 14]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"
    return wb


def get_absence_charge_code(leave_type):
    """Return workbook-style absence charge-code metadata for leave/holiday rows."""
    normalized = normalize_absence_label(leave_type)
    reference_key = LEAVE_TYPE_TO_ABSENCE_KEY.get(normalized)
    if reference_key:
        return ABSENCE_CHARGE_CODES[reference_key]

    for reference in ABSENCE_CHARGE_CODES.values():
        if normalize_absence_label(reference["name"]) == normalized:
            return reference

    clean_type = str(leave_type or "").strip()
    if clean_type:
        return {
            "code": clean_type[:3].upper(),
            "name": clean_type if "leave" in clean_type.lower() else f"{clean_type} Leave",
        }
    return ABSENCE_CHARGE_CODES["other_approved_absence"]


def get_leave_code(leave_type):
    return get_absence_charge_code(leave_type)["code"]


def get_leave_display_code(leave_type):
    normalized = normalize_absence_label(leave_type)
    return LEAVE_TYPE_DISPLAY_CODES.get(normalized, get_leave_code(leave_type))


def build_system_generated_entries(employee_id, period_start, period_end):
    """Generate locked holiday/leave entries for a timesheet period."""
    holiday_docs = list(mongo.db.holidays.find({
        "date": {"$gte": period_start, "$lte": period_end},
        "type": "company",
    }))
    holiday_entries = []
    locked_dates = {}
    public_holiday = ABSENCE_CHARGE_CODES["public_holiday"]

    for holiday in holiday_docs:
        date_key = normalize_date_key(holiday.get("date"))
        holiday_entries.append({
            "_id": ObjectId(),
            "date": date_key,
            "entry_type": "holiday",
            "holiday_name": holiday.get("name"),
            "code": public_holiday["code"],
            "display_code": "PH",
            "charge_code": public_holiday["code"],
            "charge_code_name": public_holiday["name"],
            "hours": WORKDAY_HOURS,
            "description": public_holiday["name"],
        })
        locked_dates[date_key] = {
            "kind": "holiday",
            "label": public_holiday["name"],
            "code": public_holiday["code"],
        }

    leave_docs = list(mongo.db.leaves.find({
        "employee_id": employee_id,
        "status": "Approved",
        "start_date": {"$lte": period_end},
        "end_date": {"$gte": period_start},
    }))

    leave_entries = []
    for leave in leave_docs:
        if is_early_logout_leave_type(leave.get("leave_type")):
            continue

        effective_start = normalize_date_key(leave.get("approved_start_date") or leave.get("start_date"))
        effective_end = normalize_date_key(leave.get("approved_end_date") or leave.get("end_date"))
        if not effective_start or not effective_end:
            continue
        if effective_end < period_start or effective_start > period_end:
            continue

        leave_type = leave.get("leave_type", "Leave")
        leave_reference = get_absence_charge_code(leave_type)
        leave_code = leave_reference["code"]
        leave_display_code = get_leave_display_code(leave_type)
        leave_name = leave_reference["name"]
        is_half_day = bool(leave.get("is_half_day"))
        leave_hours = WORKDAY_HOURS / 2 if is_half_day else WORKDAY_HOURS
        half_day_period = leave.get("half_day_period", "")

        for date_key in daterange_keys(max(effective_start, period_start), min(effective_end, period_end)):
            if not is_weekday_date(date_key):
                continue
            if date_key in locked_dates and locked_dates[date_key]["kind"] == "holiday":
                continue
            if date_key in locked_dates and locked_dates[date_key]["kind"] == "leave":
                continue

            leave_entries.append({
                "_id": ObjectId(),
                "date": date_key,
                "entry_type": "leave",
                "leave_type": leave_type,
                "leave_code": leave_code,
                "display_code": leave_display_code,
                "charge_code": leave_code,
                "charge_code_name": leave_name,
                "hours": leave_hours,
                "description": (
                    leave_name
                    + (f" ({half_day_period})" if is_half_day and half_day_period else "")
                ),
                "leave_id": leave["_id"],
                "is_half_day": is_half_day,
                "half_day_period": half_day_period if is_half_day else "",
            })
            if not is_half_day:
                locked_dates[date_key] = {
                    "kind": "leave",
                    "label": leave_name,
                    "code": leave_code,
                }

    return holiday_entries, leave_entries, locked_dates


def build_validated_timesheet_entries(employee_id, period_start, period_end, entries):
    """Validate user-entered work rows and merge system-generated leave/holiday rows."""
    holiday_entries, leave_entries, locked_dates = build_system_generated_entries(
        employee_id, period_start, period_end
    )

    validated_work_entries = []
    work_totals_by_date = {}
    system_hours_by_date = {}
    for item in leave_entries + holiday_entries:
        item_date = normalize_date_key(item.get("date"))
        if item_date:
            system_hours_by_date[item_date] = (
                system_hours_by_date.get(item_date, 0) + float(item.get("hours", 0) or 0)
            )

    for entry in entries or []:
        if entry.get("entry_type", "work") != "work":
            continue

        entry_date = normalize_date_key(entry.get("date"))
        if not entry_date:
            return None, None, "Entry date is required"

        locked = locked_dates.get(entry_date)
        if locked:
            return None, None, (
                f"{entry_date} is locked for {locked['label']} ({locked['code']}). "
                "Work hours cannot be entered on approved leave or holiday dates."
            )

        charge_code_id = entry.get("charge_code_id")
        if not charge_code_id:
            return None, None, f"Charge code required for work entry on {entry_date}"

        try:
            cc_obj_id = ObjectId(charge_code_id)
        except Exception:
            return None, None, f"Invalid charge_code_id on {entry_date}"

        assignment = mongo.db.charge_code_assignments.find_one({
            "employee_id": employee_id,
            "charge_code_id": cc_obj_id,
            "is_active": True,
        })
        if not assignment:
            return None, None, f"You don't have access to charge code {charge_code_id}"

        charge_code = mongo.db.charge_codes.find_one({"_id": cc_obj_id})
        work_hours = float(entry.get("hours") or 0)
        work_totals_by_date[entry_date] = work_totals_by_date.get(entry_date, 0) + work_hours

        validated_work_entries.append({
            "_id": ObjectId(),
            "date": entry_date,
            "entry_type": "work",
            "charge_code_id": cc_obj_id,
            "charge_code": charge_code.get("code") if charge_code else "Unknown",
            "charge_code_name": charge_code.get("name") if charge_code else "",
            "hours": work_hours,
            "description": entry.get("description", ""),
        })

    for entry_date, work_total in work_totals_by_date.items():
        total_for_date = work_total + system_hours_by_date.get(entry_date, 0)
        if total_for_date > WORKDAY_HOURS:
            return None, None, (
                f"{entry_date} has {total_for_date} total hours across all charge codes. "
                f"Maximum allowed is {WORKDAY_HOURS} hours."
            )

    merged_entries = validated_work_entries + leave_entries + holiday_entries
    total_hours = sum(float(item.get("hours", 0) or 0) for item in merged_entries)
    return merged_entries, total_hours, None


def get_work_hours_total(entries):
    """Return only employee-entered working hours, excluding leave and holidays."""
    return sum(
        float(entry.get("hours", 0) or 0)
        for entry in entries or []
        if entry.get("entry_type", "work") == "work"
    )


def fit_work_entries_around_system_entries(work_entries, system_entries):
    """Remove or trim work entries that now conflict with approved leave/holidays."""
    system_hours_by_date = {}
    full_day_locked_dates = set()

    for entry in system_entries or []:
        date_key = normalize_date_key(entry.get("date"))
        if not date_key:
            continue
        hours = float(entry.get("hours", 0) or 0)
        system_hours_by_date[date_key] = system_hours_by_date.get(date_key, 0) + hours
        if hours >= WORKDAY_HOURS and entry.get("entry_type") in ("leave", "holiday"):
            full_day_locked_dates.add(date_key)

    used_by_date = {}
    adjusted = []
    for entry in work_entries or []:
        if entry.get("entry_type", "work") != "work":
            continue
        date_key = normalize_date_key(entry.get("date"))
        if not date_key or date_key in full_day_locked_dates:
            continue

        original_hours = float(entry.get("hours", 0) or 0)
        remaining = WORKDAY_HOURS - system_hours_by_date.get(date_key, 0) - used_by_date.get(date_key, 0)
        if remaining <= 0:
            continue

        adjusted_hours = min(original_hours, remaining)
        if adjusted_hours <= 0:
            continue

        updated_entry = dict(entry)
        updated_entry["date"] = date_key
        updated_entry["hours"] = adjusted_hours
        used_by_date[date_key] = used_by_date.get(date_key, 0) + adjusted_hours
        adjusted.append(updated_entry)

    return adjusted


def get_fortnight_bounds_for_date(date_value):
    """Return YYYY-MM-DD fortnight bounds for the provided date."""
    if isinstance(date_value, str):
        date_value = datetime.strptime(date_value[:10], "%Y-%m-%d").date()
    elif isinstance(date_value, datetime):
        date_value = date_value.date()

    if date_value.day <= 15:
        period_start = date_value.replace(day=1)
        period_end = date_value.replace(day=15)
    else:
        period_start = date_value.replace(day=16)
        next_month = date_value.replace(day=28) + timedelta(days=4)
        period_end = next_month.replace(day=1) - timedelta(days=1)

    return (
        period_start.strftime("%Y-%m-%d"),
        period_end.strftime("%Y-%m-%d"),
    )


def iter_fortnight_periods(start_key, end_key):
    """Yield fortnight periods overlapping the supplied date range."""
    current = datetime.strptime(start_key, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_key, "%Y-%m-%d").date()
    seen = set()

    while current <= end_date:
        period_start, period_end = get_fortnight_bounds_for_date(current)
        if (period_start, period_end) not in seen:
            seen.add((period_start, period_end))
            yield period_start, period_end
        current = datetime.strptime(period_end, "%Y-%m-%d").date() + timedelta(days=1)


def create_system_generated_timesheet(employee, period_start, period_end):
    """Create a draft timesheet shell so synced leave/holiday entries are visible to employees."""
    if not employee:
        return None

    holiday_entries, leave_entries, _ = build_system_generated_entries(
        employee["_id"],
        period_start,
        period_end,
    )
    merged_entries = leave_entries + holiday_entries
    total_hours = sum(float(item.get("hours", 0) or 0) for item in merged_entries)
    now = datetime.utcnow()

    timesheet = {
        "employee_id": employee["_id"],
        "employee_name": employee.get("name", ""),
        "employee_email": employee.get("email", ""),
        "employee_department": employee.get("department", ""),
        "period_start": period_start,
        "period_end": period_end,
        "entries": merged_entries,
        "daily_overtime": {},
        "holiday_payout": {},
        "work_schedule_by_date": {},
        "employee_work_locations_by_date": {},
        "employee_assigned_locations_by_date": {},
        "total_hours": total_hours,
        "work_hours": 0,
        "status": "draft",
        "reporting_lead_id": employee.get("reportsTo"),
        "manager_id": employee.get("peopleLead"),
        "created_at": now,
        "updated_at": now,
        "system_entries_refreshed_at": now,
        "approval_history": [],
        "is_locked": False,
        "auto_created_from_leave_sync": True,
    }
    apply_employee_assignment_snapshot(timesheet, employee)
    return timesheet


def ensure_timesheets_exist_for_leave(employee_id, leave_start, leave_end):
    """Create missing fortnight timesheets so backdated approved leave appears in employee views."""
    if not employee_id or not leave_start or not leave_end:
        return 0

    employee = mongo.db.users.find_one({"_id": employee_id})
    if not employee:
        return 0

    created_count = 0
    for period_start, period_end in iter_fortnight_periods(leave_start, leave_end):
        existing = mongo.db.timesheets.find_one({
            "employee_id": employee_id,
            "period_start": period_start,
            "period_end": period_end,
        })
        if existing:
            continue

        seed_timesheet = create_system_generated_timesheet(employee, period_start, period_end)
        if not seed_timesheet:
            continue
        if not any(entry.get("entry_type") in ("leave", "holiday") for entry in seed_timesheet.get("entries", [])):
            continue

        mongo.db.timesheets.insert_one(seed_timesheet)
        created_count += 1

    if created_count:
        print(f"✅ Created {created_count} missing timesheet(s) for approved leave sync")
    return created_count


def refresh_timesheet_system_entries(timesheet):
    """Rebuild approved leave/holiday entries for an existing timesheet."""
    if not timesheet:
        return None

    period_start = timesheet.get("period_start")
    period_end = timesheet.get("period_end")
    employee_id = timesheet.get("employee_id")
    if not all([employee_id, period_start, period_end]):
        return None

    holiday_entries, leave_entries, _ = build_system_generated_entries(
        employee_id, period_start, period_end
    )
    work_entries = [
        entry for entry in timesheet.get("entries", [])
        if entry.get("entry_type", "work") == "work"
    ]
    adjusted_work_entries = fit_work_entries_around_system_entries(
        work_entries,
        leave_entries + holiday_entries,
    )
    merged_entries = adjusted_work_entries + leave_entries + holiday_entries
    total_hours = sum(float(item.get("hours", 0) or 0) for item in merged_entries)

    return {
        "entries": merged_entries,
        "total_hours": total_hours,
        "work_hours": get_work_hours_total(merged_entries),
        "updated_at": datetime.utcnow(),
        "system_entries_refreshed_at": datetime.utcnow(),
    }


def sync_timesheets_for_approved_leave(leave_record):
    """Update overlapping timesheets after a leave is approved."""
    if not leave_record or leave_record.get("status") != "Approved":
        return 0

    employee_id = leave_record.get("employee_id")
    leave_start = normalize_date_key(
        leave_record.get("approved_start_date") or leave_record.get("start_date")
    )
    leave_end = normalize_date_key(
        leave_record.get("approved_end_date") or leave_record.get("end_date")
    )
    if not employee_id or not leave_start or not leave_end:
        return 0

    created_count = ensure_timesheets_exist_for_leave(employee_id, leave_start, leave_end)
    query = {
        "employee_id": employee_id,
        "period_start": {"$lte": leave_end},
        "period_end": {"$gte": leave_start},
        "status": {"$in": ["draft", "pending_lead", "pending_manager", "approved"]},
    }
    updated_count = 0
    for timesheet in mongo.db.timesheets.find(query):
        refreshed = refresh_timesheet_system_entries(timesheet)
        if not refreshed:
            continue
        mongo.db.timesheets.update_one(
            {"_id": timesheet["_id"]},
            {"$set": refreshed},
        )
        updated_count += 1

    total_count = updated_count + created_count
    if total_count:
        print(f"✅ Synced {total_count} timesheet(s) for approved leave {leave_record.get('_id')}")
    return total_count


def sync_timesheets_for_cancelled_leave(leave_record):
    """Refresh overlapping timesheets after an approved leave is cancelled."""
    if not leave_record:
        return 0

    employee_id = leave_record.get("employee_id")
    leave_start = normalize_date_key(
        leave_record.get("approved_start_date") or leave_record.get("start_date")
    )
    leave_end = normalize_date_key(
        leave_record.get("approved_end_date") or leave_record.get("end_date")
    )
    if not employee_id or not leave_start or not leave_end:
        return 0

    query = {
        "employee_id": employee_id,
        "period_start": {"$lte": leave_end},
        "period_end": {"$gte": leave_start},
        "status": {"$in": ["draft", "pending_lead", "pending_manager", "approved"]},
    }
    updated_count = 0
    for timesheet in mongo.db.timesheets.find(query):
        refreshed = refresh_timesheet_system_entries(timesheet)
        if not refreshed:
            continue
        mongo.db.timesheets.update_one(
            {"_id": timesheet["_id"]},
            {"$set": refreshed},
        )
        updated_count += 1

    if updated_count:
        print(f"✅ Synced {updated_count} timesheet(s) for cancelled leave {leave_record.get('_id')}")
    return updated_count


def refresh_timesheet_for_read(timesheet):
    """Refresh system entries before returning a timesheet to the UI."""
    refreshed = refresh_timesheet_system_entries(timesheet)
    if not refreshed:
        return apply_timesheet_workflow_preference(timesheet)

    mongo.db.timesheets.update_one(
        {"_id": timesheet["_id"]},
        {"$set": refreshed},
    )
    timesheet.update(refreshed)
    return apply_timesheet_workflow_preference(timesheet)


# ========================================
# CREATE / SUBMIT TIMESHEET
# ========================================

@timesheet_bp.route("/create", methods=["POST"])
def create_timesheet():
    try:
        data = request.get_json()
        employee_id  = data.get("employee_id")
        period_start = data.get("period_start")
        period_end   = data.get("period_end")
        entries      = data.get("entries", [])

        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "Missing required fields: employee_id, period_start, period_end"}), 400

        limit_error = validate_daily_work_hours(entries)
        if limit_error:
            return jsonify({"error": limit_error}), 400

        daily_overtime, adjustment_error = normalize_daily_adjustments(
            data.get("daily_overtime", {}), period_start, period_end, "Daily overtime"
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        holiday_payout, adjustment_error = normalize_daily_adjustments(
            data.get("holiday_payout", {}), period_start, period_end, "Holiday payout"
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_schedule_by_date, adjustment_error = normalize_work_schedule(
            data.get("work_schedule_by_date", data.get("work_schedule", {})),
            period_start,
            period_end,
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_locations_by_date = normalize_location_map(
            data.get("employee_work_locations_by_date")
            or data.get("work_locations_by_date")
            or data.get("daily_locations"),
            period_start,
            period_end,
        )
        assigned_locations_by_date = normalize_location_map(
            data.get("employee_assigned_locations_by_date")
            or data.get("assigned_locations_by_date"),
            period_start,
            period_end,
        )

        try:
            emp_obj_id = ObjectId(employee_id)
        except Exception:
            return jsonify({"error": "Invalid employee_id format"}), 400

        employee = mongo.db.users.find_one({"_id": emp_obj_id})
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        existing = mongo.db.timesheets.find_one({
            "employee_id": emp_obj_id,
            "period_start": period_start,
            "period_end":   period_end,
        })
        if is_fortnight_entry_blocked(emp_obj_id, period_start, period_end):
            return jsonify({
                "error": get_fortnight_deadline_label(period_start, period_end, employee_id=emp_obj_id)
            }), 400
        if existing and existing.get("status") == "approved" and not is_approved_edit_window_open(existing):
            return jsonify({
                "error": "This timesheet has already been approved and is locked."
            }), 400

        reporting_lead_id = employee.get("reportsTo")
        if not reporting_lead_id:
            return jsonify({"error": "No reporting lead found for employee"}), 404

        validated_entries, total_hours, entry_error = build_validated_timesheet_entries(
            emp_obj_id, period_start, period_end, entries
        )
        if entry_error:
            return jsonify({"error": entry_error}), 400

        now = datetime.utcnow()
        timesheet = {
            "employee_id":         emp_obj_id,
            "employee_name":       employee.get("name"),
            "employee_email":      employee.get("email"),
            "employee_department": employee.get("department", ""),
            "period_start":        period_start,
            "period_end":          period_end,
            "entries":             validated_entries,
            "daily_overtime":      daily_overtime,
            "holiday_payout":      holiday_payout,
            "work_schedule_by_date": work_schedule_by_date,
            "employee_work_locations_by_date": work_locations_by_date,
            "employee_assigned_locations_by_date": assigned_locations_by_date,
            "total_hours":         total_hours,
            "work_hours":          get_work_hours_total(validated_entries),
            "status":              "pending_lead",
            "reporting_lead_id":   reporting_lead_id,
            "manager_id":          employee.get("peopleLead"),
            "created_at":          now,
            "updated_at":          now,
            "submitted_at":        now,
            "approval_history":    [],
            "escalation_level":    0,
            "escalation_history":  [],
        }
        apply_employee_assignment_snapshot(timesheet, employee)
        apply_timesheet_workflow_preference(timesheet, employee=employee)
        current_approver = get_timesheet_current_approver(timesheet, employee=employee)
        if not current_approver:
            return jsonify({"error": "No approver found for employee"}), 404
        timesheet["current_approver_id"] = current_approver["_id"]

        if existing:
            if existing.get("status") == "approved":
                timesheet["reopened_from_approved"] = True
                timesheet["approved_edit_window_used_at"] = now
                timesheet["is_locked"] = False
            mongo.db.timesheets.update_one({"_id": existing["_id"]}, {"$set": timesheet})
            timesheet_id = existing["_id"]
        else:
            result = mongo.db.timesheets.insert_one(timesheet)
            timesheet_id = result.inserted_id
        timesheet["_id"] = timesheet_id

        # Notify the reporting lead
        msg = (
            f"{employee.get('name')} submitted a timesheet for "
            f"{period_start} to {period_end} ({total_hours}h)"
        )
        notify_timesheet_pending_approver(
            timesheet,
            current_approver["_id"],
            event="submitted",
            message=msg,
        )
        queue_timesheet_submission_emails(timesheet, employee=employee)

        return jsonify({
            "message": "Timesheet submitted successfully",
            "timesheet_id": str(timesheet_id),
            "total_hours": total_hours,
        }), 201

    except Exception as e:
        print(f"❌ Error creating timesheet: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ========================================
# UPDATE TIMESHEET ENTRIES (draft / recalled)
# ========================================

@timesheet_bp.route("/update/<timesheet_id>", methods=["PUT"])
def update_timesheet(timesheet_id):
    """Update entries on a draft or recalled timesheet before re-submitting."""
    try:
        data = request.get_json()
        entries = data.get("entries", [])

        limit_error = validate_daily_work_hours(entries)
        if limit_error:
            return jsonify({"error": limit_error}), 400

        ts = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not ts:
            return jsonify({"error": "Timesheet not found"}), 404
        if not is_employee_editable_timesheet(ts):
            return jsonify({"error": "Only draft, rejected, or eligible approved timesheets can be updated"}), 400

        daily_overtime, adjustment_error = normalize_daily_adjustments(
            data.get("daily_overtime", ts.get("daily_overtime", {})),
            ts.get("period_start"),
            ts.get("period_end"),
            "Daily overtime",
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        holiday_payout, adjustment_error = normalize_daily_adjustments(
            data.get("holiday_payout", ts.get("holiday_payout", {})),
            ts.get("period_start"),
            ts.get("period_end"),
            "Holiday payout",
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_schedule_by_date, adjustment_error = normalize_work_schedule(
            data.get("work_schedule_by_date", data.get("work_schedule", ts.get("work_schedule_by_date", {}))),
            ts.get("period_start"),
            ts.get("period_end"),
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        validated_entries, total_hours, entry_error = build_validated_timesheet_entries(
            ts["employee_id"],
            ts.get("period_start"),
            ts.get("period_end"),
            entries,
        )
        if entry_error:
            return jsonify({"error": entry_error}), 400

        update_data = {
            "entries":        validated_entries,
            "daily_overtime": daily_overtime,
            "holiday_payout": holiday_payout,
            "work_schedule_by_date": work_schedule_by_date,
            "total_hours":    total_hours,
            "work_hours":     get_work_hours_total(validated_entries),
            "updated_at":     datetime.utcnow(),
        }
        work_location_keys = (
            "employee_work_locations_by_date",
            "work_locations_by_date",
            "daily_locations",
        )
        if any(key in data for key in work_location_keys):
            update_data["employee_work_locations_by_date"] = normalize_location_map(
                data.get("employee_work_locations_by_date")
                or data.get("work_locations_by_date")
                or data.get("daily_locations"),
                ts.get("period_start"),
                ts.get("period_end"),
            )
        if "employee_assigned_locations_by_date" in data or "assigned_locations_by_date" in data:
            update_data["employee_assigned_locations_by_date"] = normalize_location_map(
                data.get("employee_assigned_locations_by_date")
                or data.get("assigned_locations_by_date"),
                ts.get("period_start"),
                ts.get("period_end"),
            )

        employee = mongo.db.users.find_one({"_id": ts["employee_id"]})
        if employee:
            apply_employee_assignment_snapshot(update_data, employee)

        update_data = build_resubmission_update(ts, update_data)

        mongo.db.timesheets.update_one(
            {"_id": ObjectId(timesheet_id)},
            {"$set": update_data}
        )
        return jsonify({"message": "Timesheet updated", "total_hours": total_hours}), 200

    except Exception as e:
        print(f"❌ Error updating timesheet: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/save_draft", methods=["POST"])
def save_timesheet_draft():
    """Create or update a draft timesheet without submitting it for approval."""
    try:
        data = request.get_json() or {}
        employee_id = data.get("employee_id")
        period_start = data.get("period_start")
        period_end = data.get("period_end")
        entries = data.get("entries", [])

        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "Missing required fields: employee_id, period_start, period_end"}), 400

        limit_error = validate_daily_work_hours(entries)
        if limit_error:
            return jsonify({"error": limit_error}), 400

        daily_overtime, adjustment_error = normalize_daily_adjustments(
            data.get("daily_overtime", {}), period_start, period_end, "Daily overtime"
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        holiday_payout, adjustment_error = normalize_daily_adjustments(
            data.get("holiday_payout", {}), period_start, period_end, "Holiday payout"
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_schedule_by_date, adjustment_error = normalize_work_schedule(
            data.get("work_schedule_by_date", data.get("work_schedule", {})),
            period_start,
            period_end,
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_locations_by_date = normalize_location_map(
            data.get("employee_work_locations_by_date")
            or data.get("work_locations_by_date")
            or data.get("daily_locations"),
            period_start,
            period_end,
        )
        assigned_locations_by_date = normalize_location_map(
            data.get("employee_assigned_locations_by_date")
            or data.get("assigned_locations_by_date"),
            period_start,
            period_end,
        )

        try:
            emp_obj_id = ObjectId(employee_id)
        except Exception:
            return jsonify({"error": "Invalid employee_id format"}), 400

        employee = mongo.db.users.find_one({"_id": emp_obj_id})
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        existing = mongo.db.timesheets.find_one({
            "employee_id": emp_obj_id,
            "period_start": period_start,
            "period_end": period_end,
        })
        if is_fortnight_entry_blocked(emp_obj_id, period_start, period_end):
            return jsonify({"error": get_fortnight_deadline_label(period_start, period_end, employee_id=emp_obj_id)}), 400
        if existing and existing.get("status") in ("pending_lead", "pending_manager"):
            return jsonify({"error": "Submitted or approved timesheets cannot be saved as drafts"}), 400
        if existing and existing.get("status") == "approved" and not is_approved_edit_window_open(existing):
            return jsonify({"error": "Approved timesheets can only be edited during the configured correction window"}), 400
        if existing and existing.get("status") == "draft" and existing.get("reopened_from_approved") and not is_correction_window_open_for_timesheet(existing):
            return jsonify({"error": "The correction window has closed for this previously approved timesheet"}), 400

        validated_entries, total_hours, entry_error = build_validated_timesheet_entries(
            emp_obj_id, period_start, period_end, entries
        )
        if entry_error:
            return jsonify({"error": entry_error}), 400
        now = datetime.utcnow()
        draft = {
            "employee_id": emp_obj_id,
            "employee_name": employee.get("name"),
            "employee_email": employee.get("email"),
            "employee_department": employee.get("department", ""),
            "period_start": period_start,
            "period_end": period_end,
            "entries": validated_entries,
            "daily_overtime": daily_overtime,
            "holiday_payout": holiday_payout,
            "work_schedule_by_date": work_schedule_by_date,
            "employee_work_locations_by_date": work_locations_by_date,
            "employee_assigned_locations_by_date": assigned_locations_by_date,
            "total_hours": total_hours,
            "work_hours": get_work_hours_total(validated_entries),
            "status": "draft",
            "reporting_lead_id": employee.get("reportsTo"),
            "manager_id": employee.get("peopleLead"),
            "updated_at": now,
            "approval_history": existing.get("approval_history", []) if existing else [],
        }
        if existing and existing.get("status") == "approved":
            draft["reopened_from_approved"] = True
            draft["approved_edit_window_used_at"] = now
            draft["is_locked"] = False
        apply_employee_assignment_snapshot(draft, employee)
        apply_timesheet_workflow_preference(draft, employee=employee)

        if existing:
            mongo.db.timesheets.update_one({"_id": existing["_id"]}, {"$set": draft})
            timesheet_id = existing["_id"]
            clear_timesheet_reminder_state(timesheet_id)
        else:
            draft["created_at"] = now
            result = mongo.db.timesheets.insert_one(draft)
            timesheet_id = result.inserted_id

        return jsonify({
            "message": "Draft saved successfully",
            "timesheet_id": str(timesheet_id),
            "total_hours": total_hours,
        }), 200

    except Exception as e:
        print(f"❌ Error saving draft: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/delete/<timesheet_id>", methods=["DELETE"])
def delete_timesheet(timesheet_id):
    """Delete an editable timesheet."""
    try:
        ts = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not ts:
            return jsonify({"error": "Timesheet not found"}), 404
        if ts.get("status") in ("pending_lead", "pending_manager"):
            return jsonify({"error": "Submitted or approved timesheets cannot be deleted"}), 400
        if ts.get("reopened_from_approved"):
            return jsonify({"error": "Previously approved timesheets cannot be deleted"}), 400

        mongo.db.timesheets.delete_one({"_id": ObjectId(timesheet_id)})
        return jsonify({"message": "Timesheet deleted successfully"}), 200

    except Exception as e:
        print(f"❌ Error deleting timesheet: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# SUBMIT EXISTING TIMESHEET
# ========================================

@timesheet_bp.route("/submit/<timesheet_id>", methods=["PUT"])
def submit_timesheet(timesheet_id):
    """Submit a draft/rejected timesheet into the approval queue."""
    try:
        ts = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not ts:
            return jsonify({"error": "Timesheet not found"}), 404
        if is_fortnight_entry_blocked(ts.get("employee_id"), ts.get("period_start"), ts.get("period_end")):
            return jsonify({"error": get_fortnight_deadline_label(ts.get("period_start"), ts.get("period_end"), employee_id=ts.get("employee_id"))}), 400
        if ts.get("status") == "approved" and not is_approved_edit_window_open(ts):
            return jsonify({"error": "Timesheet is already approved"}), 400
        if ts.get("status") == "draft" and ts.get("reopened_from_approved") and not is_correction_window_open_for_timesheet(ts):
            return jsonify({"error": "The correction window has closed for this previously approved timesheet"}), 400

        limit_error = validate_daily_work_hours(ts.get("entries", []))
        if limit_error:
            return jsonify({"error": limit_error}), 400

        daily_overtime, adjustment_error = normalize_daily_adjustments(
            ts.get("daily_overtime", {}),
            ts.get("period_start"),
            ts.get("period_end"),
            "Daily overtime",
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        holiday_payout, adjustment_error = normalize_daily_adjustments(
            ts.get("holiday_payout", {}),
            ts.get("period_start"),
            ts.get("period_end"),
            "Holiday payout",
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_schedule_by_date, adjustment_error = normalize_work_schedule(
            ts.get("work_schedule_by_date", {}),
            ts.get("period_start"),
            ts.get("period_end"),
        )
        if adjustment_error:
            return jsonify({"error": adjustment_error}), 400

        work_entries = [
            entry for entry in (ts.get("entries") or [])
            if entry.get("entry_type", "work") == "work"
        ]
        validated_entries, total_hours, entry_error = build_validated_timesheet_entries(
            ts["employee_id"],
            ts.get("period_start"),
            ts.get("period_end"),
            work_entries,
        )
        if entry_error:
            return jsonify({"error": entry_error}), 400

        now = datetime.utcnow()
        update_data = {
            "entries":       validated_entries,
            "daily_overtime": daily_overtime,
            "holiday_payout": holiday_payout,
            "work_schedule_by_date": work_schedule_by_date,
            "total_hours":   total_hours,
            "work_hours":    get_work_hours_total(validated_entries),
            "status":       "pending_lead",
            "submitted_at": now,
            "updated_at":   now,
            "is_locked":    False,
            "escalation_level": 0,
            "escalated_on": None,
            "previous_approver_id": None,
        }
        if ts.get("status") == "approved" or ts.get("reopened_from_approved") or ts.get("requires_reapproval"):
            update_data["resubmitted_after_approval_at"] = now
            update_data["requires_reapproval"] = True
        employee = mongo.db.users.find_one({"_id": ts["employee_id"]})
        if employee:
            apply_employee_assignment_snapshot(update_data, employee)
        working_copy = dict(ts)
        working_copy.update(update_data)
        apply_timesheet_workflow_preference(working_copy, employee=employee)
        update_data["reporting_lead_id"] = working_copy.get("reporting_lead_id")
        if "reporting_lead_name" in working_copy:
            update_data["reporting_lead_name"] = working_copy.get("reporting_lead_name", "")
        if "reporting_lead_email" in working_copy:
            update_data["reporting_lead_email"] = working_copy.get("reporting_lead_email", "")
        if "timesheet_workflow_preferences" in working_copy:
            update_data["timesheet_workflow_preferences"] = working_copy.get("timesheet_workflow_preferences")
        current_approver = get_timesheet_current_approver(working_copy, employee=employee)
        if not current_approver:
            return jsonify({"error": "No approver found for this timesheet"}), 404
        update_data["current_approver_id"] = current_approver["_id"]

        mongo.db.timesheets.update_one(
            {"_id": ObjectId(timesheet_id)},
            {"$set": update_data}
        )

        # Notify reporting lead
        employee_name = ts.get("employee_name", "An employee")
        is_resubmission = bool(ts.get("reopened_from_approved") or ts.get("requires_reapproval") or ts.get("lead_approved_at"))
        event_name = "resubmitted" if is_resubmission else "submitted"
        notify_timesheet_pending_approver(
            {**working_copy, "_id": ObjectId(timesheet_id)},
            current_approver["_id"],
            event=event_name,
            message=(
                f"{employee_name} submitted a timesheet for "
                f"{ts.get('period_start')} to {ts.get('period_end')}"
            ),
        )
        working_copy["_id"] = ObjectId(timesheet_id)
        queue_timesheet_submission_emails(working_copy, employee=employee)

        return jsonify({
            "message":     "Timesheet submitted",
            "total_hours": total_hours,
        }), 200

    except Exception as e:
        print(f"❌ Error submitting timesheet: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# RECALL TIMESHEET (Employee retracts for editing)
# ========================================

@timesheet_bp.route("/recall/<timesheet_id>", methods=["PUT"])
def recall_timesheet(timesheet_id):
    try:
        ts = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not ts:
            return jsonify({"error": "Timesheet not found"}), 404

        if ts.get("status") == "approved":
            return jsonify({"error": "Cannot recall an approved timesheet"}), 400

        mongo.db.timesheets.update_one(
            {"_id": ObjectId(timesheet_id)},
            {"$set": {
                "status":     "draft",
                "updated_at": datetime.utcnow(),
            }}
        )
        clear_timesheet_reminder_state(timesheet_id)
        print(f"✅ Timesheet {timesheet_id} recalled to draft")
        return jsonify({"message": "Timesheet recalled to draft"}), 200

    except Exception as e:
        print(f"❌ Error recalling timesheet: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# GET EMPLOYEE'S TIMESHEETS
# ========================================

@timesheet_bp.route("/employee/<employee_id>", methods=["GET"])
def get_employee_timesheets(employee_id):
    try:
        if not employee_id or len(employee_id) != 24:
            return jsonify({"error": "Invalid employee_id format"}), 400

        timesheets = list(
            mongo.db.timesheets.find({"employee_id": ObjectId(employee_id)}).sort("period_start", -1)
        )
        timesheets = [refresh_timesheet_for_read(ts) for ts in timesheets]
        timesheets = [enrich_timesheet_with_employee_assignments(ts) for ts in timesheets]
        timesheets = [annotate_timesheet_editability(ts) for ts in timesheets]
        return jsonify(serialize_all(timesheets)), 200

    except Exception as e:
        print(f"❌ Error fetching employee timesheets: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# GET PENDING TIMESHEETS FOR LEAD
# ========================================

@timesheet_bp.route("/pending/lead/<user_id>", methods=["GET"])
def get_pending_for_lead(user_id):
    try:
        requester = resolve_requester()
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        query = {"status": {"$in": ["pending_lead", "pending_manager"]}}
        if has_admin_menu_access(requester, "timesheets"):
            pass
        elif str(requester.get("_id")) == str(user_id):
            query["current_approver_id"] = ObjectId(user_id)
        else:
            return jsonify({"error": "You do not have permission to view these approvals"}), 403

        timesheets = list(mongo.db.timesheets.find(query).sort("submitted_at", -1))
        timesheets = [refresh_timesheet_for_read(ts) for ts in timesheets]
        timesheets = [enrich_timesheet_with_employee_assignments(ts) for ts in timesheets]
        timesheets = [annotate_timesheet_editability(ts) for ts in timesheets]
        return jsonify(serialize_all(timesheets)), 200

    except Exception as e:
        print(f"❌ Error fetching pending timesheets for lead: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# GET PENDING TIMESHEETS FOR MANAGER
# KEPT for backwards compatibility — always returns empty list now
# ========================================

@timesheet_bp.route("/pending/manager/<user_id>", methods=["GET"])
def get_pending_for_manager(user_id):
    try:
        requester = resolve_requester()
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        query = {"status": "pending_manager"}
        if has_admin_menu_access(requester, "timesheets"):
            pass
        elif str(requester.get("_id")) == str(user_id):
            query["current_approver_id"] = ObjectId(user_id)
        else:
            return jsonify({"error": "You do not have permission to view these approvals"}), 403

        timesheets = list(mongo.db.timesheets.find(query).sort("submitted_at", -1))
        timesheets = [refresh_timesheet_for_read(ts) for ts in timesheets]
        timesheets = [enrich_timesheet_with_employee_assignments(ts) for ts in timesheets]
        timesheets = [annotate_timesheet_editability(ts) for ts in timesheets]
        return jsonify(serialize_all(timesheets)), 200
    except Exception as e:
        print(f"❌ Error fetching pending timesheets for manager: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# LEAD APPROVAL — grants full approval
# ========================================

@timesheet_bp.route("/approve/lead/<timesheet_id>", methods=["PUT"])
def lead_approve_timesheet(timesheet_id):
    try:
        data        = request.get_json()
        comments    = data.get("comments", "")
        approver_name_input = str(data.get("approver_name") or "").strip()
        requester   = resolve_requester()

        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        timesheet = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not timesheet:
            return jsonify({"error": "Timesheet not found"}), 404
        if timesheet.get("status") not in ("pending_lead", "pending_manager"):
            return jsonify({"error": "Timesheet is not pending approval"}), 400

        requester_id = str(requester.get("_id"))
        current_approver_id = str(timesheet.get("current_approver_id") or "")
        is_timesheet_admin = has_admin_menu_access(requester, "timesheets")
        if not (is_timesheet_admin or requester_id == current_approver_id):
            return jsonify({"error": "You do not have permission to approve this timesheet"}), 403
        if is_timesheet_admin and not approver_name_input:
            return jsonify({"error": "approver_name is required for admin approval"}), 400

        approver = requester
        approver_id = approver.get("_id")
        approver_name = approver_name_input or approver.get("name") or ""
        employee = mongo.db.users.find_one({"_id": timesheet.get("employee_id")})
        # Lead approval is always final — only look for a next approver when
        # the timesheet is already at the manager stage (pending_manager).
        if is_timesheet_admin or timesheet.get("status") == "pending_lead":
            next_approver = None
        else:
            next_approver = get_next_timesheet_approver(timesheet, employee=employee)

        approval_entry = {
            "stage":         "lead" if timesheet.get("status") == "pending_lead" else "manager",
            "action":        "approved",
            "approver_id":   approver_id,
            "approver_name": approver_name,
            "comments":      comments,
            "timestamp":     datetime.utcnow(),
        }

        if next_approver:
            move_timesheet_to_next_approver(timesheet, next_approver, approval_entry, escalation=False)
            print(f"✅ Timesheet {timesheet_id} advanced to next approver {next_approver.get('name')}")
            return jsonify({"message": "Timesheet approved and forwarded"}), 200

        final_stage_label = "lead" if approval_entry["stage"] == "lead" else "manager"
        finalize_timesheet_approval(timesheet, approver_id, approver_name, approval_entry, final_stage_label=final_stage_label)
        print(f"✅ Timesheet {timesheet_id} fully approved by {final_stage_label} {approver_name}")
        return jsonify({"message": "Timesheet approved"}), 200

    except Exception as e:
        print(f"❌ Error in lead approval: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ========================================
# LEAD REJECTION
# ========================================

@timesheet_bp.route("/reject/lead/<timesheet_id>", methods=["PUT"])
def lead_reject_timesheet(timesheet_id):
    try:
        data             = request.get_json()
        rejection_reason = data.get("rejection_reason", "").strip()
        approver_name_input = str(data.get("approver_name") or "").strip()
        requester        = resolve_requester()

        if not rejection_reason:
            return jsonify({"error": "Rejection reason is required"}), 400
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        timesheet = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not timesheet:
            return jsonify({"error": "Timesheet not found"}), 404
        if timesheet.get("status") not in ("pending_lead", "pending_manager"):
            return jsonify({"error": "Timesheet is not pending approval"}), 400

        requester_id = str(requester.get("_id"))
        current_approver_id = str(timesheet.get("current_approver_id") or "")
        is_timesheet_admin = has_admin_menu_access(requester, "timesheets")
        if not (is_timesheet_admin or requester_id == current_approver_id):
            return jsonify({"error": "You do not have permission to reject this timesheet"}), 403
        if is_timesheet_admin and not approver_name_input:
            return jsonify({"error": "approver_name is required for admin rejection"}), 400

        rejector = requester
        rejector_id = rejector.get("_id")
        rejector_name = approver_name_input or rejector.get("name") or ""
        stage = "lead" if timesheet.get("status") == "pending_lead" else "manager"
        rejected_status = "rejected_by_lead" if stage == "lead" else "rejected_by_manager"

        rejection_entry = {
            "stage":         stage,
            "action":        "rejected",
            "approver_id":   rejector_id,
            "approver_name": rejector_name,
            "comments":      rejection_reason,
            "timestamp":     datetime.utcnow(),
        }

        mongo.db.timesheets.update_one(
            {"_id": ObjectId(timesheet_id)},
            {
                "$set": {
                    "status":           rejected_status,
                    "lead_rejected_at": datetime.utcnow() if stage == "lead" else timesheet.get("lead_rejected_at"),
                    "lead_rejected_by": rejector_name if stage == "lead" else timesheet.get("lead_rejected_by"),
                    "lead_rejected_by_id": rejector_id if stage == "lead" else timesheet.get("lead_rejected_by_id"),
                    "manager_rejected_at": datetime.utcnow() if stage == "manager" else timesheet.get("manager_rejected_at"),
                    "manager_rejected_by": rejector_name if stage == "manager" else timesheet.get("manager_rejected_by"),
                    "manager_rejected_by_id": rejector_id if stage == "manager" else timesheet.get("manager_rejected_by_id"),
                    "current_approver_id": None,
                    "rejection_reason": rejection_reason,
                    "updated_at":       datetime.utcnow(),
                },
                "$push": {"approval_history": rejection_entry},
            },
        )

        rejection_target = build_timesheet_notification_target(timesheet, view="entry")
        rejection_message = (
            f"Your timesheet ({timesheet.get('period_start')} to {timesheet.get('period_end')}) "
            f"was rejected by your {stage}. Reason: {rejection_reason}"
        )
        create_notification(
            user_id=timesheet["employee_id"],
            notification_type="timesheet_rejected",
            message=rejection_message,
            related_timesheet_id=timesheet_id,
            target=rejection_target,
            reminder_group=f"timesheet:{timesheet_id}:rejected:{timesheet['employee_id']}",
        )
        set_timesheet_reminder_state(
            timesheet_id,
            event="rejected",
            mode="until_read",
            recipient_id=timesheet["employee_id"],
            target=rejection_target,
            message=rejection_message,
        )

        return jsonify({"message": f"Timesheet rejected by {stage}"}), 200

    except Exception as e:
        print(f"❌ Error in lead rejection: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ========================================
# MANAGER APPROVAL — KEPT for backwards compatibility only
# ========================================

@timesheet_bp.route("/approve/manager/<timesheet_id>", methods=["PUT"])
def manager_approve_timesheet(timesheet_id):
    try:
        data = request.get_json()
        comments = data.get("comments", "")
        approver_name_input = str(data.get("approver_name") or "").strip()
        requester = resolve_requester()
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        timesheet = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not timesheet:
            return jsonify({"error": "Timesheet not found"}), 404
        if timesheet.get("status") != "pending_manager":
            return jsonify({"error": "Timesheet is not pending manager approval"}), 400

        requester_id = str(requester.get("_id"))
        current_approver_id = str(timesheet.get("current_approver_id") or "")
        is_timesheet_admin = has_admin_menu_access(requester, "timesheets")
        if not (is_timesheet_admin or requester_id == current_approver_id):
            return jsonify({"error": "You do not have permission to approve this timesheet"}), 403
        if is_timesheet_admin and not approver_name_input:
            return jsonify({"error": "approver_name is required for admin approval"}), 400

        approver = requester
        approver_id = approver.get("_id")
        approver_name = approver_name_input or approver.get("name") or ""
        employee = mongo.db.users.find_one({"_id": timesheet.get("employee_id")})
        next_approver = None if is_timesheet_admin else get_next_timesheet_approver(timesheet, employee=employee)

        approval_entry = {
            "stage":         "manager",
            "action":        "approved",
            "approver_id":   approver_id,
            "approver_name": approver_name,
            "comments":      comments,
            "timestamp":     datetime.utcnow(),
        }

        if next_approver:
            move_timesheet_to_next_approver(timesheet, next_approver, approval_entry, escalation=False)
            return jsonify({"message": "Timesheet approved and forwarded"}), 200

        finalize_timesheet_approval(timesheet, approver_id, approver_name, approval_entry, final_stage_label="manager")

        return jsonify({"message": "Timesheet fully approved"}), 200

    except Exception as e:
        print(f"❌ Error in manager approval: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ========================================
# MANAGER REJECTION — KEPT for backwards compatibility only
# ========================================

@timesheet_bp.route("/reject/manager/<timesheet_id>", methods=["PUT"])
def manager_reject_timesheet(timesheet_id):
    try:
        data             = request.get_json()
        rejection_reason = data.get("rejection_reason", "").strip()
        approver_name_input = str(data.get("approver_name") or "").strip()
        requester = resolve_requester()

        if not rejection_reason:
            return jsonify({"error": "Rejection reason is required"}), 400
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        timesheet = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not timesheet:
            return jsonify({"error": "Timesheet not found"}), 404
        if timesheet.get("status") != "pending_manager":
            return jsonify({"error": "Timesheet is not pending manager approval"}), 400

        requester_id = str(requester.get("_id"))
        current_approver_id = str(timesheet.get("current_approver_id") or "")
        is_timesheet_admin = has_admin_menu_access(requester, "timesheets")
        if not (is_timesheet_admin or requester_id == current_approver_id):
            return jsonify({"error": "You do not have permission to reject this timesheet"}), 403
        if is_timesheet_admin and not approver_name_input:
            return jsonify({"error": "approver_name is required for admin rejection"}), 400

        rejector = requester
        rejector_id = rejector.get("_id")
        rejector_name = approver_name_input or rejector.get("name") or ""

        rejection_entry = {
            "stage":         "manager",
            "action":        "rejected",
            "approver_id":   rejector_id,
            "approver_name": rejector_name,
            "comments":      rejection_reason,
            "timestamp":     datetime.utcnow(),
        }

        mongo.db.timesheets.update_one(
            {"_id": ObjectId(timesheet_id)},
            {
                "$set": {
                    "status":              "rejected_by_manager",
                    "manager_rejected_at": datetime.utcnow(),
                    "manager_rejected_by": rejector_name,
                    "manager_rejected_by_id": rejector_id,
                    "current_approver_id": None,
                    "rejection_reason":    rejection_reason,
                    "updated_at":          datetime.utcnow(),
                },
                "$push": {"approval_history": rejection_entry},
            },
        )

        manager_rejection_target = build_timesheet_notification_target(timesheet, view="entry")
        manager_rejection_message = (
            f"Your timesheet ({timesheet.get('period_start')} to {timesheet.get('period_end')}) "
            f"was rejected by manager. Reason: {rejection_reason}"
        )
        create_notification(
            user_id=timesheet["employee_id"],
            notification_type="timesheet_rejected",
            message=manager_rejection_message,
            related_timesheet_id=timesheet_id,
            target=manager_rejection_target,
            reminder_group=f"timesheet:{timesheet_id}:rejected:{timesheet['employee_id']}",
        )
        set_timesheet_reminder_state(
            timesheet_id,
            event="rejected",
            mode="until_read",
            recipient_id=timesheet["employee_id"],
            target=manager_rejection_target,
            message=manager_rejection_message,
        )

        return jsonify({"message": "Timesheet rejected by manager"}), 200

    except Exception as e:
        print(f"❌ Error in manager rejection: {str(e)}")
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ========================================
# GET ALL TIMESHEETS (ADMIN)
# ========================================

@timesheet_bp.route("/all", methods=["GET"])
def get_all_timesheets():
    try:
        _, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        timesheets = list(mongo.db.timesheets.find().sort("submitted_at", -1))

        result = []
        for ts in timesheets:
            ts = refresh_timesheet_for_read(ts)
            ts = enrich_timesheet_with_employee_assignments(ts)
            ts = annotate_timesheet_editability(ts)
            ts = serialize_all(ts)
            employee_id = ts.get("employee_id")
            if employee_id and (not ts.get("employee_name") or not ts.get("employee_department")):
                try:
                    emp = mongo.db.users.find_one({"_id": ObjectId(employee_id)})
                    if emp:
                        ts["employee_name"]       = ts.get("employee_name") or emp.get("name", "")
                        ts["employee_email"]      = ts.get("employee_email") or emp.get("email", "")
                        ts["employee_department"] = emp.get("department", "")
                except Exception:
                    pass

            reporting_lead_id = ts.get("reporting_lead_id")
            if reporting_lead_id and (not ts.get("reporting_lead_name") or not ts.get("reporting_lead_email")):
                try:
                    reporting_lead = mongo.db.users.find_one({"_id": ObjectId(reporting_lead_id)})
                    if reporting_lead:
                        ts["reporting_lead_name"] = reporting_lead.get("name", "")
                        ts["reporting_lead_email"] = reporting_lead.get("email", "")
                except Exception:
                    pass
            result.append(ts)

        return jsonify(result), 200

    except Exception as e:
        print(f"❌ Error in /all: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/preferences", methods=["GET"])
def get_timesheet_preferences():
    try:
        _, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        employee_id = request.args.get("employee_id", "").strip()
        period_start = request.args.get("period_start", "").strip()
        period_end = request.args.get("period_end", "").strip()
        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "employee_id, period_start, and period_end are required"}), 400

        preference = get_timesheet_preference(ObjectId(employee_id), period_start, period_end)
        if not preference:
            return jsonify({
                "employee_id": employee_id,
                "period_start": period_start,
                "period_end": period_end,
                "reviewers": [],
                "notifications": [],
                "delegates": [],
                "approvers": [],
            }), 200

        return jsonify(serialize_all(preference)), 200
    except Exception as e:
        print(f"❌ Error fetching timesheet preferences: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/period-access", methods=["GET"])
def get_timesheet_period_access():
    try:
        requester = resolve_requester()
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401

        employee_id = request.args.get("employee_id", "").strip()
        period_start = request.args.get("period_start", "").strip()
        period_end = request.args.get("period_end", "").strip()
        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "employee_id, period_start, and period_end are required"}), 400

        employee_obj_id = ObjectId(employee_id)
        if not (has_admin_menu_access(requester, "timesheets") or str(requester.get("_id")) == str(employee_obj_id)):
            return jsonify({"error": "You do not have permission to view this fortnight"}), 403

        access_state = get_period_access_state(employee_obj_id, period_start, period_end)
        unlock_record = access_state.pop("unlock_record", None)
        payload = {
            **access_state,
            "employee_id": str(employee_obj_id),
        }
        if unlock_record:
            payload["unlock_record"] = serialize_all(unlock_record)
        return jsonify(payload), 200
    except Exception as e:
        print(f"❌ Error fetching period access: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/block-settings", methods=["GET"])
def get_timesheet_block_settings_route():
    try:
        requester, error_response = require_admin()
        if error_response:
            return error_response

        global_settings = get_timesheet_block_settings()
        overrides = list(mongo.db.timesheet_block_overrides.find({"is_active": True}).sort("updated_at", -1))
        history = list(mongo.db.timesheet_block_history.find().sort("changed_at", -1).limit(50))
        return jsonify({
            **global_settings,
            "global": global_settings,
            "employee_overrides": [serialize_timesheet_block_override(item) for item in overrides],
            "history": [serialize_timesheet_block_history(item) for item in history],
        }), 200
    except Exception as e:
        print(f"❌ Error fetching timesheet block settings: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/block-settings", methods=["PUT"])
def update_timesheet_block_settings_route():
    try:
        requester, error_response = require_admin()
        if error_response:
            return error_response

        data = request.get_json() or {}
        settings = normalize_timesheet_block_settings(data)
        now = datetime.utcnow()
        mongo.db.system_settings.update_one(
            {"key": TIMESHEET_BLOCK_SETTINGS_KEY},
            {
                "$set": {
                    "key": TIMESHEET_BLOCK_SETTINGS_KEY,
                    "value": settings,
                    "updated_at": now,
                    "updated_by": requester.get("_id"),
                    "updated_by_name": requester.get("name", ""),
                    "updated_by_email": requester.get("email", ""),
                },
                "$setOnInsert": {"created_at": now},
            },
            upsert=True,
        )
        record_timesheet_block_history("updated", "global", settings, requester)
        overrides = list(mongo.db.timesheet_block_overrides.find({"is_active": True}).sort("updated_at", -1))
        history = list(mongo.db.timesheet_block_history.find().sort("changed_at", -1).limit(50))
        return jsonify({
            **settings,
            "global": settings,
            "employee_overrides": [serialize_timesheet_block_override(item) for item in overrides],
            "history": [serialize_timesheet_block_history(item) for item in history],
        }), 200
    except Exception as e:
        print(f"❌ Error updating timesheet block settings: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/block-settings/employee-overrides", methods=["POST"])
def save_timesheet_employee_block_override():
    try:
        requester, error_response = require_admin()
        if error_response:
            return error_response

        data = request.get_json() or {}
        employee_id = str(data.get("employee_id", "")).strip()
        if not employee_id:
            return jsonify({"error": "employee_id is required"}), 400

        employee = mongo.db.users.find_one({"_id": ObjectId(employee_id)})
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        settings = normalize_timesheet_block_settings(data)
        notes = str(data.get("notes", "")).strip()
        now = datetime.utcnow()
        override_doc = {
            "employee_id": employee["_id"],
            "employee_name": employee.get("name", ""),
            "employee_email": employee.get("email", ""),
            **settings,
            "notes": notes,
            "is_active": True,
            "updated_at": now,
            "updated_by": requester.get("_id"),
            "updated_by_name": requester.get("name", ""),
            "updated_by_email": requester.get("email", ""),
        }

        override_id = str(data.get("_id") or "").strip()
        if override_id:
            mongo.db.timesheet_block_overrides.update_one(
                {"_id": ObjectId(override_id)},
                {"$set": override_doc},
            )
            saved = mongo.db.timesheet_block_overrides.find_one({"_id": ObjectId(override_id)})
            action = "updated"
        else:
            override_doc["created_at"] = now
            result = mongo.db.timesheet_block_overrides.insert_one(override_doc)
            saved = mongo.db.timesheet_block_overrides.find_one({"_id": result.inserted_id})
            action = "created"

        record_timesheet_block_history(action, "employee", settings, requester, employee=employee, notes=notes)
        history = list(mongo.db.timesheet_block_history.find().sort("changed_at", -1).limit(50))
        return jsonify({
            "override": serialize_timesheet_block_override(saved),
            "history": [serialize_timesheet_block_history(item) for item in history],
        }), 200
    except Exception as e:
        print(f"❌ Error saving employee timesheet block override: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/block-settings/employee-overrides/<override_id>", methods=["DELETE"])
def delete_timesheet_employee_block_override(override_id):
    try:
        requester, error_response = require_admin()
        if error_response:
            return error_response

        existing = mongo.db.timesheet_block_overrides.find_one({"_id": ObjectId(override_id)})
        mongo.db.timesheet_block_overrides.update_one(
            {"_id": ObjectId(override_id)},
            {
                "$set": {
                    "is_active": False,
                    "deleted_at": datetime.utcnow(),
                    "deleted_by": requester.get("_id"),
                }
            },
        )
        if existing:
            record_timesheet_block_history(
                "removed",
                "employee",
                existing,
                requester,
                employee={
                    "_id": existing.get("employee_id"),
                    "name": existing.get("employee_name", ""),
                    "email": existing.get("employee_email", ""),
                },
                notes=existing.get("notes", ""),
            )
        history = list(mongo.db.timesheet_block_history.find().sort("changed_at", -1).limit(50))
        return jsonify({
            "message": "Employee-specific block rule removed",
            "history": [serialize_timesheet_block_history(item) for item in history],
        }), 200
    except Exception as e:
        print(f"❌ Error deleting employee timesheet block override: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/period-unlocks", methods=["PUT"])
def save_timesheet_period_unlock():
    try:
        requester = resolve_requester()
        if not requester:
            return jsonify({"error": "A valid requester is required"}), 401
        if requester.get("role") != "Admin":
            return jsonify({"error": "Only Admin users can manage blocked fortnights"}), 403

        data = request.get_json() or {}
        employee_id = str(data.get("employee_id", "")).strip()
        period_start = str(data.get("period_start", "")).strip()
        period_end = str(data.get("period_end", "")).strip()
        unlocked = bool(data.get("unlocked"))
        notes = str(data.get("notes", "")).strip()
        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "employee_id, period_start, and period_end are required"}), 400

        employee = mongo.db.users.find_one({"_id": ObjectId(employee_id)})
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        update_doc = {
            "employee_id": employee["_id"],
            "employee_name": employee.get("name", ""),
            "employee_email": employee.get("email", ""),
            "period_start": period_start,
            "period_end": period_end,
            "is_active": unlocked,
            "notes": notes,
            "updated_at": datetime.utcnow(),
            "updated_by_admin_id": requester.get("_id"),
            "updated_by_admin_name": requester.get("name", ""),
        }
        if unlocked:
            update_doc["unlocked_at"] = datetime.utcnow()
        else:
            update_doc["relocked_at"] = datetime.utcnow()

        mongo.db.timesheet_period_unlocks.update_one(
            {
                "employee_id": employee["_id"],
                "period_start": period_start,
                "period_end": period_end,
            },
            {"$set": update_doc, "$setOnInsert": {"created_at": datetime.utcnow()}},
            upsert=True,
        )

        access_state = get_period_access_state(employee["_id"], period_start, period_end)
        unlock_record = access_state.pop("unlock_record", None)
        payload = {
            **access_state,
            "employee_id": str(employee["_id"]),
        }
        if unlock_record:
            payload["unlock_record"] = serialize_all(unlock_record)
        return jsonify(payload), 200
    except Exception as e:
        print(f"❌ Error saving period unlock: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/preferences", methods=["PUT"])
def save_timesheet_preferences():
    try:
        requester, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        data = request.get_json() or {}
        employee_id = data.get("employee_id", "")
        period_start = str(data.get("period_start", "")).strip()
        period_end = str(data.get("period_end", "")).strip()
        if not all([employee_id, period_start, period_end]):
            return jsonify({"error": "employee_id, period_start, and period_end are required"}), 400

        employee = mongo.db.users.find_one({"_id": ObjectId(employee_id)})
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        document = build_timesheet_preference_document(
            employee["_id"],
            period_start,
            period_end,
            data,
            managed_by=requester,
        )
        mongo.db.timesheet_preferences.update_one(
            {
                "employee_id": employee["_id"],
                "period_start": period_start,
                "period_end": period_end,
            },
            {"$set": document, "$setOnInsert": {"created_at": datetime.utcnow()}},
            upsert=True,
        )
        saved = get_timesheet_preference(employee["_id"], period_start, period_end)
        return jsonify(serialize_all(saved)), 200
    except Exception as e:
        print(f"❌ Error saving timesheet preferences: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/reports/lop-summary", methods=["GET"])
def get_lop_summary_report():
    try:
        _, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        today = datetime.utcnow()
        start_dt = parse_report_date(request.args.get("start_date"), today.replace(day=1))
        end_dt = parse_report_date(request.args.get("end_date"), today)
        if start_dt > end_dt:
            return jsonify({"error": "start_date cannot be after end_date"}), 400

        start_key = start_dt.strftime("%Y-%m-%d")
        end_key = end_dt.strftime("%Y-%m-%d")
        result = build_lop_report_rows(start_key, end_key, {
            "department": request.args.get("department"),
            "employee_status": request.args.get("employee_status"),
            "search": request.args.get("search"),
        })
        return jsonify({
            "company": "Naxrita Solutions Private Limited",
            "title": "Loss Of Pay Summary",
            "start_date": start_key,
            "end_date": end_key,
            **result,
        }), 200
    except Exception as e:
        print(f"❌ Error in /reports/lop-summary: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/reports/lop-summary/filters", methods=["GET"])
def get_lop_summary_filters():
    try:
        _, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        employees = list(mongo.db.users.find({"role": {"$ne": "Admin"}}).sort("name", 1))
        departments = sorted({emp.get("department") for emp in employees if emp.get("department")})
        return jsonify({
            "departments": departments,
            "employees": [
                {
                    "value": " ".join([
                        str(emp.get("employeeId", "")),
                        str(emp.get("name", "")),
                        str(emp.get("email", "")),
                    ]).strip(),
                    "label": emp.get("name", "") or emp.get("email", "") or emp.get("employeeId", ""),
                    "description": " · ".join(
                        item for item in [
                            emp.get("employeeId", ""),
                            emp.get("email", ""),
                            emp.get("department", ""),
                        ] if item
                    ),
                }
                for emp in employees
            ],
        }), 200
    except Exception as e:
        print(f"❌ Error in /reports/lop-summary/filters: {str(e)}")
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/reports/lop-summary/export", methods=["GET"])
def export_lop_summary_report():
    try:
        _, error_response = require_admin_menu_access("timesheets")
        if error_response:
            return error_response

        today = datetime.utcnow()
        start_dt = parse_report_date(request.args.get("start_date"), today.replace(day=1))
        end_dt = parse_report_date(request.args.get("end_date"), today)
        if start_dt > end_dt:
            return jsonify({"error": "start_date cannot be after end_date"}), 400

        start_key = start_dt.strftime("%Y-%m-%d")
        end_key = end_dt.strftime("%Y-%m-%d")
        result = build_lop_report_rows(start_key, end_key, {
            "department": request.args.get("department"),
            "employee_status": request.args.get("employee_status"),
            "search": request.args.get("search"),
        })
        workbook = create_lop_report_workbook(start_key, end_key, result["rows"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"lop_summary_{start_key}_to_{end_key}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"❌ Error exporting LOP summary: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# AUTO-POPULATE HOLIDAYS FOR PERIOD
# ========================================

@timesheet_bp.route("/populate_holidays", methods=["POST"])
def populate_holidays():
    try:
        data         = request.get_json()
        period_start = data.get("period_start")
        period_end   = data.get("period_end")

        if not period_start or not period_end:
            return jsonify({"error": "period_start and period_end are required"}), 400

        holidays = list(mongo.db.holidays.find({
            "date": {"$gte": period_start, "$lte": period_end},
            "type": "company",
        }))

        public_holiday = ABSENCE_CHARGE_CODES["public_holiday"]
        holiday_entries = [
            {
                "date":         h["date"],
                "entry_type":   "holiday",
                "holiday_name": h.get("name"),
                "hours":        WORKDAY_HOURS,
                "code":         public_holiday["code"],
                "display_code": "PH",
                "charge_code":  public_holiday["code"],
                "charge_code_name": public_holiday["name"],
                "description":  public_holiday["name"],
            }
            for h in holidays
        ]

        return jsonify({"holidays": holiday_entries, "count": len(holiday_entries)}), 200

    except Exception as e:
        print(f"❌ Error populating holidays: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# GET TEAM TIMESHEETS
# ========================================

@timesheet_bp.route("/team/<manager_email>", methods=["GET"])
def get_team_timesheets(manager_email):
    try:
        manager = mongo.db.users.find_one({"email": manager_email})
        if not manager:
            return jsonify([]), 200

        employees = list(mongo.db.users.find({
            "reportsTo": manager["_id"],
            "role":      {"$ne": "Admin"},
        }))

        all_ts = []
        for emp in employees:
            ts_list = list(mongo.db.timesheets.find({"employee_id": emp["_id"]}).sort("period_start", -1))
            all_ts.extend(ts_list)

        all_ts = [refresh_timesheet_for_read(ts) for ts in all_ts]
        all_ts = [enrich_timesheet_with_employee_assignments(ts) for ts in all_ts]
        return jsonify(serialize_all(all_ts)), 200

    except Exception as e:
        print(f"❌ Error fetching team timesheets: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ========================================
# GET SINGLE TIMESHEET BY ID
# ========================================

@timesheet_bp.route("/<timesheet_id>", methods=["GET"])
def get_timesheet(timesheet_id):
    try:
        ts = mongo.db.timesheets.find_one({"_id": ObjectId(timesheet_id)})
        if not ts:
            return jsonify({"error": "Timesheet not found"}), 404
        ts = refresh_timesheet_for_read(ts)
        ts = enrich_timesheet_with_employee_assignments(ts)
        ts = annotate_timesheet_editability(ts)
        return jsonify(serialize_all(ts)), 200

    except Exception as e:
        print(f"❌ Error fetching timesheet: {str(e)}")
        return jsonify({"error": str(e)}), 500
